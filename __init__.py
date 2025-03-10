from flask import Flask, render_template, url_for,request,redirect,session,jsonify, send_file
import os,sys,stat
from werkzeug.utils import secure_filename
import Customer , Listing,Reviews,Report,operatoractions,Feedback #classes
from Delivery import Delivery
from Forms import CustomerSignupForm, ReportForm, CustomerLoginForm, ListingForm, ReviewForm, CustomerUpdateForm, \
    SearchBar, OperatorLoginForm, OperatorLoginVerifyForm, SearchUserField, OperatorSuspendUser, OperatorTerminateUser, \
    OperatorRestoreUser, DeliveryForm, SearchListingIDField,SearchListingStatusField,SearchTransactionField   # our forms
from Forms import OperatorDisableListing,OperatorRestoreListing,SearchListingField,SearchReportField,SearchOperatorActionField,FeedbackForm,FilterForm,UpdateFeedback,ReplyFeedback,SearchUserStatus,FilterFeedback,FilterTransactions
import Operatorstats
import Email,Search,Notifications
import shelve, Customer
from pathlib import Path
from Messages import User
import string
import random
from datetime import datetime
import Filters
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import uuid
import re
import io
from webscraping_ebay_amazon import get_ebay_estimated_price
from chat import get_response
app = Flask(__name__)



def check_allowed_file(filename):
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def check_upload_file_type(file,type,id): #file obj here, listing/userid here too, type too
    file_to_upload = file
    if type == "listing" and check_allowed_file(file.filename):
        file.filename = f"listing{id}.jpg"
        check_dupe_file(file_to_upload,"listingpics")
   
    elif type == "customer" and check_allowed_file(file.filename):
        file.filename = f"customer{id}.jpg"
        check_dupe_file(file_to_upload,"profilepics")
    elif type == "message" and check_allowed_file(file.filename):
        file.filename = f"message{id}.jpg"
        check_dupe_file(file_to_upload,"messagepics")
    else:
        print("invalid type or ID")

def check_dupe_file(file,type):
    myfile = Path(f"static/{type}/{file.filename}")
    if myfile.is_file():
        os.remove(f"static/{type}/{file.filename}")
        file_uploaded = file
        folder = type
        upload_file(folder,file_uploaded)

    else:
        file_uploaded = file
        folder = type
        upload_file(folder,file_uploaded)

def upload_file(folder,file):
    filename = secure_filename(file.filename)
    file.save(os.path.join(f"static/{folder}", filename))
    #for some reason it keeps returning ERRNO13 aka no permission but err it works so idk

def get_searchquery(formdict,outputlist):
    print(formdict)
    conditions = ['category1','category2','category3','category4','category5','condition_barelyused','condition_frequentlyused','condition_wornout','sortlatest']
    for condition in conditions:
        if formdict.get(condition) == True:
            outputlist.append(condition)
        else:
            pass
    return outputlist

def get_matchinglistingID(searchquerylist,outputlist):
    dbmain = shelve.open('main.db','c') #note that redirect urls convert anything and everything into a string, so use session attributes to store lists such as this one
    listings_dict = {}
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #first sort the search query into 2 sections, category, conditions
    conditionfilter = []
    categoryfilter = []
    filtercategory = False
    filtercondition = False
    for condition in searchquerylist:
        print(f"Condition is {condition}")
        if "category" in condition:
            categoryfilter.append(condition)
            filtercategory = True
        if "condition" in condition:
            conditionfilter.append(condition)
            filtercondition = True
    filtersortlatest = False
    for condition in searchquerylist:
        if "sortlatest" in condition:
            filtersortlatest = True
    
    #filter category
    if filtercategory  == True: #2 scenarios, A) the outputlist is empty , B) it is not. If A) this becomes a simple if match then add. If B) If match add, if no match delete
        for key in listings_dict: #loop through objects
            listing = listings_dict.get(key)
            for category in categoryfilter:
                
                if category == "category1":
                    Filters.category(listing,'Category 1',outputlist,"addonly")
                if category == "category2":
                    Filters.category(listing,'Category 2',outputlist,"addonly")
                if category == "category3":
                    Filters.category(listing,'Category 3',outputlist,"addonly")
                if category == "category4":
                    Filters.category(listing,'Category 4',outputlist,"addonly")
                if category == "category5":
                    Filters.category(listing,'Category 5',outputlist,"addonly")

    #filter condition
    # 2 scenarios here
    print(outputlist)
    if filtercondition == True:
        if filtercategory == True:#implement remove if not match and to read directly from outputlist as it shld be not emtpy due to previous filter
            checklist = []
            for ID in outputlist: #so basically if does not match, it will remove, if does it adds
                if ID not in checklist:
                    checklist.append(ID)
                    listing = listings_dict.get(ID)
                    for condition in conditionfilter:
                        if condition == "condition_barelyused":
                            print("b_used")
                            Filters.condition(listing,'Barely used',outputlist,"addanddelete")
                        if condition == "condition_frequentlyused":
                            print("f_used")
                            Filters.condition(listing,'Frequently used',outputlist,"addanddelete")
                        if condition == "condition_wornout":
                            print("ud_used")
                            Filters.condition(listing,'Worn out',outputlist,'addanddelete')
                    print("looped")
                    
        if filtercategory == False: #outputlist will be empty
            for key in listings_dict:
                listing = listings_dict.get(key)
                for condition in conditionfilter:
                    if condition == "condition_barelyused":
                        Filters.condition(listing,'Barely used',outputlist,"addonly")
                    if condition == "condition_frequentlyused":
                        Filters.condition(listing,'Frequently used',outputlist,"addonly")
                    if condition == "condition_wornout":
                        Filters.condition(listing,'Worn out',outputlist,'addonly')
                print("L")
    #sort latest
    if filtersortlatest == True:
        if outputlist == []:
            for key in listings_dict:
                listing = listings_dict.get(key)
                outputlist.append(listing.get_ID())
                outputlist.sort(reverse=True)
        else:
            outputlist.sort(reverse=True)

    return outputlist 

def deduper(inputlistID):
    return list(dict.fromkeys(inputlistID))

def ID_to_obj(inputlistID,outputlistobj):
    dbmain = shelve.open('main.db','c')
    listings_dict = {}
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    for ID in inputlistID:
        listing = listings_dict.get(ID)
        if Search.check_listing(listing):
            outputlistobj.append(listing)#for now it will be to ID, easier during testing
        else:
            pass
    return outputlistobj

def filterdict(dict):
    if True in dict.values():
        return True
    else:
        return False
    

def send_welcomenotifcation(id): #id of person to add notifactions to
    dbmain = shelve.open('main.db','c')
    
    customers_dict = {}
    notifications_dict = {}
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening customer in main.db")
    try:
        if "Notifications" in dbmain:
            notifications_dict = dbmain["Notifications"] #sync local with db1
        else:
            dbmain['Notifications'] = notifications_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    
    try:
        dbmain = shelve.open('main.db','c')    
        Notifications.Notifications.count_ID = dbmain["NotificationsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Notifications count or count is at 0")
    
    #create notifcation obj
    notification = Notifications.Notifications(id,"Welcome to Freesell!")
    notifications_dict[notification.get_ID()] = notification
    dbmain["Notifications"] = notifications_dict
    dbmain["NotificationsCount"] = Notifications.Notifications.count_ID
    #add it to customer notifcations
    customer = customers_dict.get(id)
    customer.add_notifications(notification.get_ID()) #adds notifs
    dbmain["Customers"] = customers_dict

    #send the notifcation to their email

    Email.send_signup_notification_gmail(customer.get_email(),customer.get_username())


#current_sessionID IS FOR THE PROFILE!!
#Current session ID
#MAKE SURE AN USER WITH ID "X" EXISTS!      
session_ID = 0

#makes creating customer page easier, PS:NOT ACTUAL PAGE!
def template():
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    global session_ID
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    #search func
    try:
        
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data)) #get the word from the search field
    except:
        pass
    
    #get notifs
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        customer_notifications = customer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if request.method == "POST" and filterform.validate():
            searchconditionlist = []
            get_searchquery(filterform.data,searchconditionlist)
            session['filters'] = searchconditionlist
            return redirect(url_for('filterresults'))
    except:
        pass
    #get username for navbar
    customer = customers_dict.get(session_ID)
    current_username = customer.get_username()
    return render_template('template.html', current_sessionID = session_ID,searchform =search_field,customer_notifications = customer_notifications,filterform = filterform,current_username=current_username)

#create op stats obj here
@app.route('/', methods = ['GET', 'POST']) #shld be the same as href for buttons,links,navbar, etc...
def Customerhome():
    dbmain = shelve.open('main.db','c')
    listings_dict = {}
    customers_dict = {} #local one
    operatorstats_dict = {}
    global session_ID
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    #search func
    try:
        
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data)) #get the word from the search field
    except:
        pass

    try:
        if "Operatorstats" in dbmain:
            operatorstats_dict = dbmain["Operatorstats"]  # sync local with db1
        else:
            dbmain['Operatorstats'] = operatorstats_dict # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    #opstats code
    try:
        dbmain = shelve.open('main.db', 'c')
        Operatorstats.Operatorstats.count_ID = dbmain["Operatorstatscount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Operatorstats count or count is at 0")
    if Operatorstats.Operatorstats.count_ID == 0:
        operatorstats = Operatorstats.Operatorstats()
        operatorstats_dict[operatorstats.get_ID()] = operatorstats
        dbmain['Operatorstats'] = operatorstats_dict
        dbmain['Operatorstatscount'] = Operatorstats.Operatorstats.count_ID
    else:
        pass
    print(operatorstats_dict)


    #get notifs
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        customer_notifications = customer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if request.method == "POST" and filterform.validate():
            searchconditionlist = []
            get_searchquery(filterform.data,searchconditionlist)
            session['filters'] = searchconditionlist
            return redirect(url_for('filterresults'))
    except:
        pass

    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"

    #data analytics for users
    customer = customers_dict.get(session_ID)
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"]
        else:
            dbmain["Listings"] = listings_dict
    except:
        print("Error in opening Listings data.")
    return render_template('Customerhome.html', current_sessionID = session_ID,searchform =search_field,customer_notifications = customer_notifications,filterform = filterform,customer= customer)


@app.route('/suspended')
def Customersuspended_terminatedhome():
    dbmain = shelve.open('main.db','c')
    customers_dict = {} 
    global session_ID
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    customer = customers_dict.get(int(session_ID))
    return render_template('Customersuspend_terminate.html', current_sessionID = session_ID, customer = customer)

@app.post("/predict")
def predict():
    text = request.get_json().get("message")
    response = get_response(text)
    message = {"answer": response}
    return jsonify(message)

#listingalgo
@app.route('/profile/<int:id>', methods = ['GET', 'POST'])
def Customerprofile(id):
    global session_ID
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    listings_dict = {}
    reports_dict = {}
    report_form = ReportForm(request.form)
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)

    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    #sync listing dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Listing.Listing.count_ID = dbmain["ListingsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Listing count or count is at 0")

    #sync report dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Reports" in dbmain:
            reports_dict = dbmain["Reports"] #sync local with db2
        else:
            dbmain['Reports'] = reports_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Report.Report.count_ID = dbmain["ReportsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Report count or count is at 0")


    #code for profile pic img 
    pfpimg = os.path.join('..\static','profilepics')
    user_id = os.path.join(pfpimg,'hermos.jpg') 
    idreal = id
    print(idreal)
    #get ID list of current user listings

    customer = customers_dict.get(idreal)
    customer_listings = customer.get_listings()

    listing_list = []
    if current_customer.get_status() == "active":
        for key in listings_dict:
            print(key)
            if key in customer_listings:
                listing = listings_dict.get(key)
                if session_ID != idreal: #if you are NOT viewing your own profile
                    if Search.check_listing(listing):
                        listing_list.append(listing)
                    else:
                        pass
                else:
                    listing_list.append(listing)

    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    ##get username for navbar
    if session_ID != 0:

        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"

    return render_template('Customerprofile.html',customer_imgid = user_id, customer=customer, current_customer = current_customer,
                            current_sessionID = session_ID,listings_list = listing_list,form=report_form,searchform =search_field,customer_notifications = customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/updateprofile/<int:id>', methods = ['GET', 'POST'])
def updateCustomerprofile(id):
    global session_ID
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)

    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")
    customer = customers_dict.get(id) #get current customer
    customer_update_form = CustomerUpdateForm(request.form,username=customer.get_username(),email = customer.get_email(),password = customer.get_password())
    
    if request.method == 'POST' and customer_update_form.validate():
        print(customer_update_form.username.data, customer_update_form.email.data,customer_update_form.password.data)
        customer = customers_dict.get(id) #get current customer
        if customer_update_form.confirmpassword.data == customer.get_password():

            if customer_update_form.email.data != "":
                customer.set_email(customer_update_form.email.data)
            else:
                print(customer_update_form.email.data)

            if customer_update_form.username.data !="":
                customer.set_username(customer_update_form.username.data)
            else:
                print(customer_update_form.username.data)

            if customer_update_form.password.data !="":
                customer.set_password(customer_update_form.password.data)
            else:
                print(customer_update_form.password.data)

            dbmain['Customers'] = customers_dict
            dbmain.close()
            #upload img
            file = request.files['file']
            check_upload_file_type(file,"customer",customer.get_id())
            return redirect(url_for('Customerprofile', id = id))
        else:
            print("Error in changing profile info")
            return redirect(url_for('Customerprofile', id = id))
        
    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0    

    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    
    ##get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"

    return render_template("CustomerUpdateProfile.html",current_sessionID = session_ID,form=customer_update_form,searchform =search_field,customer_notifications = customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/profilereviews/<int:id>', methods = ['GET', 'POST'])
def Customerprofile_reviews(id):
    global session_ID
    dbmain = shelve.open('main.db','c')
    customers_dict = {}
    reviews_dict ={}
    reports_dict = {}
    report_form = ReportForm(request.form)
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)

    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Reviews" in dbmain:
            reviews_dict = dbmain["Reviews"] #sync local with db1
        else:
            dbmain['Reviews'] = reviews_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Reviews.Reviews.count_ID = dbmain["ReviewsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Review count or count is at 0")

    #sync report dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Reports" in dbmain:
            reports_dict = dbmain["Reports"] #sync local with db2
        else:
            dbmain['Reports'] = reports_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Report.Report.count_ID = dbmain["ReportsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Report count or count is at 0")


    
    current_customer = customers_dict.get(id)
    customer_reviews = current_customer.get_reviews()#return list of review IDs
    print(customer_reviews)
    customer_reviews_list = [] #THIS is the one sent to the html 
    
    for key in reviews_dict:
        print(key)
        if key in customer_reviews:
            print(key)
            review = reviews_dict.get(key)
            customer_reviews_list.append(review)
    #report function
    if request.method == 'POST' and report_form.validate():
        current_customer = customers_dict.get(id)

        #create report obj and store it
        print(report_form.category.data,report_form.report_text.data)
        report = Report.Report(session_ID,id,current_customer.get_username(),report_form.category.data,report_form.report_text.data)
        reports_dict[report.get_ID()] = report #store obj in dict
        dbmain['Reports'] = reports_dict
        dbmain['ReportsCount'] = Report.Report.count_ID

        #store report in offender's report_listings
        current_customer = customers_dict.get(id)
        current_customer.add_reports(report.get_ID())
        dbmain['Customers'] = customers_dict
        dbmain.close()
        redirect(url_for('Customerprofile', id=id))
        #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0   
    
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"
    return render_template('Customerprofile_reviews.html',current_customer = current_customer, customer = customer ,number_of_reviews = len(customer_reviews_list), list_reviews = customer_reviews_list, current_sessionID = session_ID,form=report_form,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/usernametaken')
def usernametaken():
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    global session_ID
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    #search func
    try:
        
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data)) #get the word from the search field
    except:
        pass
    
    #get notifs
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        customer_notifications = customer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if request.method == "POST" and filterform.validate():
            searchconditionlist = []
            get_searchquery(filterform.data,searchconditionlist)
            session['filters'] = searchconditionlist
            return redirect(url_for('filterresults'))
    except:
        pass
    current_username = "nil"
    return render_template('Customersignupinvalidusername.html', current_sessionID = session_ID,searchform =search_field,customer_notifications = customer_notifications,filterform = filterform,current_username=current_username)

@app.route('/emailtaken')
def emailtaken():
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    global session_ID
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    #search func
    try:
        
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data)) #get the word from the search field
    except:
        pass
    
    #get notifs
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        customer_notifications = customer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if request.method == "POST" and filterform.validate():
            searchconditionlist = []
            get_searchquery(filterform.data,searchconditionlist)
            session['filters'] = searchconditionlist
            return redirect(url_for('filterresults'))
    except:
        pass
    current_username = "nil"
    return render_template('Customersignupinvalidemail.html', current_sessionID = session_ID,searchform =search_field,customer_notifications = customer_notifications,filterform = filterform,current_username=current_username)

#opstatshere- user
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    global session_ID
    session_ID = 0
    create_customer_form = CustomerSignupForm(request.form)
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    dbmain = shelve.open('main.db','c')  
    customers_dict = {} #local one
    customers_username_list = []
    customers_email_list = []
    if request.method == 'POST' and create_customer_form.validate():

        try:
            if "Customers" in dbmain:
                customers_dict = dbmain["Customers"] 
            else:
                dbmain['Customers'] = customers_dict
        except:
            print("Error in opening main.db")

        #makes checking thru doing account creation a bit easier
        try:
            if "Customer_usernames" in dbmain:
                customers_username_list = dbmain["Customer_usernames"]
            else:
                dbmain['Customer_usernames'] = customers_username_list
        except:
            print("Error in opening main.db")

        #makes checking thru doing account creation a bit easier
        try:
            if "Customer_emails" in dbmain:
                customers_email_list = dbmain["Customer_emails"]
            else:
                dbmain['Customer_emails'] = customers_email_list
        except:
            print("Error in opening main.db")
        #sync IDs
        try:
            dbmain = shelve.open('main.db','c')    
            Customer.Customer.count_id = dbmain["CustomerCount"]
        except:
            print("Error in retrieving data from DB main Customer count or count is at 0")

        #checks if username is taken
        if create_customer_form.username.data in customers_username_list:
            print("Error! Username is already in db!")
            return(redirect(url_for('usernametaken')))
        elif create_customer_form.email.data in customers_email_list:
            print("Error! Email is already in db!")
            return(redirect(url_for('emailtaken')))
        else:
            #creation of customer obj
            customer =  Customer.Customer(create_customer_form.username.data.strip(), create_customer_form.email.data,create_customer_form.password.data.strip())
            customers_dict[customer.get_id()] = customer
            dbmain['Customers'] = customers_dict
            dbmain['CustomerCount'] = Customer.Customer.count_id
            #update opstats
            try:
                Operatorstats.operatorstats_users("total","plus")
                Operatorstats.operatorstats_users("active","plus")
            except:
                print("Error! Operator stats did not update")
            
            #stores username and email into local var
            customers_username_list.append(str(customer.get_username()))
            customers_email_list.append(str(customer.get_email()))
            dbmain['Customer_usernames'] = customers_username_list
            dbmain['Customer_emails'] = customers_email_list
            #verifies new user is stored
            customers_dict = dbmain['Customers'] #sync local dict with db1
            customer = customers_dict[customer.get_id()]
            session_ID = Customer.Customer.count_id
            #upload img
            file = request.files['file']
            check_upload_file_type(file,"customer",customer.get_id())
            dbmain.close() #sync the count as it updated when creating the object, if you want to hard reset the count, add a line in customer class to hard reset it to 0 so when syncing, db's one becomes 0
            #notifs
            send_welcomenotifcation(customer.get_id())

            return redirect(url_for('Customerhome'))
        

    
    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"

    return render_template("CustomerSignup.html",form=create_customer_form,current_sessionID = session_ID,searchform =search_field,customer_notifications = customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/login', methods = ['GET', 'POST'])
def login():
    global session_ID
    session_ID = 0
    login_customer_form = CustomerLoginForm(request.form)
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
                customers_dict = dbmain["Customers"] #sync local with db1
        else:
                dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")
    if request.method == 'POST' and login_customer_form.validate() and Customer.Customer.count_id ==0:
        return redirect(url_for("signup"))
    else:
        pass 
    if request.method == 'POST' and login_customer_form.validate() and Customer.Customer.count_id !=0 :
        
        #retrieve data from the form
        input_username = login_customer_form.username.data.strip()
        input_password = login_customer_form.password.data.strip()

        #check

        for key in customers_dict:
            customer = customers_dict[key] #ID
            if input_username == customer.get_username():
                if input_password == customer.get_password():
                    session_ID = key #current session = this ID
                    break
                else:
                    print("password verification failed")
            else:
                pass
        
        print(f"\n*start of message*Login success, current session ID is {session_ID}\n*end of message*")
        if customer.get_status() == "active":

            return redirect(url_for('Customerhome'))

        else:
            return redirect(url_for('Customersuspended_terminatedhome'))
    else:
        pass

    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    
    ##get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"
    return render_template("CustomerLogin.html",form=login_customer_form,current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username = current_username)

@app.route('/loginoptions',methods = ['GET', 'POST'])
def loginoptions():
    global session_ID
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    session_ID = 0
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    reports_dict = {}
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    #sync report dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Reports" in dbmain:
            reports_dict = dbmain["Reports"] #sync local with db2
        else:
            dbmain['Reports'] = reports_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"

    return render_template('Login.html',current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

#opstatshere - listing
@app.route('/createlisting', methods = ['GET', 'POST'])
def createlisting():
    global session_ID
    dbmain = shelve.open('main.db','c')
    listings_dict = {}
    customers_dict = {}
    create_listing_form = ListingForm(request.form)
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #UPLOAD FOLDER, should be local
    UPLOAD_FOLDER = 'static/listingpics/'
    app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
    #sync listing dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")

    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Listing.Listing.count_ID = dbmain["ListingsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Listing count or count is at 0")

     #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")


    #retrieve data from form
    print("code reached here")
    if request.method == 'POST' and create_listing_form.validate():
        current_date = datetime.now()
        formatted_date = current_date.strftime('%d/%m/%y')
        customer = customers_dict.get(session_ID)
        listing = Listing.Listing(session_ID,customer.get_username(),create_listing_form.title.data,create_listing_form.description.data,create_listing_form.condition.data,create_listing_form.category.data, formatted_date, None)
        listings_dict[listing.get_ID()] = listing
        dbmain['Listings'] = listings_dict
        dbmain['ListingsCount'] = Listing.Listing.count_ID
        try:
            Operatorstats.operatorstats_listings("total","plus")
            Operatorstats.operatorstats_listings("available","plus")
        except:
            print("Error! Operator stats did not update.")
        #upload img
        file = request.files['file']
        check_upload_file_type(file,"listing",listing.get_ID())

        for key in customers_dict:
            if key == session_ID:
                customer = customers_dict[key]
                customer.add_listings(Listing.Listing.count_ID)
                dbmain['Customers'] = customers_dict #syncs with db1
                break #stop the for loop if this fulfills
        return redirect(url_for('Customerprofile', id=session_ID))# returns to YOUR profile
    
    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"
    return render_template('CustomerCreateListing.html', form = create_listing_form, current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/updateListing/<int:id>/', methods=['GET', 'POST'])
def updateListing(id):
    global session_ID
    filterform = FilterForm(request.form)
    search_field = SearchBar(request.form)
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    listings_dict = {}
    #retreive listing info
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    listing = listings_dict.get(id)
    update_listing_form = ListingForm(request.form,category = listing.get_category(),condition = listing.get_condition(),title = listing.get_title(),description = listing.get_description())
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    
    if request.method == 'POST' and update_listing_form.validate():
        #upload img
        file = request.files['file']
        check_upload_file_type(file,"listing",listing.get_ID())
        listing.set_title(update_listing_form.title.data)
        listing.set_category(update_listing_form.category.data)
        listing.set_description(update_listing_form.description.data)
        listing.set_condition(update_listing_form.condition.data)
        dbmain['Listings'] = listings_dict #sync local to db2
        dbmain.close() 
        return redirect(url_for('Customerprofile', id = session_ID)) #go back to profile page after submit
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"
    return render_template('CustomerUpdateListing.html', form = update_listing_form,current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,listing = listing,current_username=current_username) #to render the form 


@app.route('/viewListing/<int:id>/', methods=['GET', 'POST'])
def viewListing(id):
    global session_ID
    dbmain = shelve.open('main.db', 'c')
    listings_dict = {}
    customers_dict = {}
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"]  # sync local with db2
        else:
            dbmain['Listings'] = listings_dict  # sync db2 with local (basically null)
    except:
        print("Error in opening main.db")

    # PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"]  # sync local with db1
        else:
            dbmain['Customers'] = customers_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    listing = listings_dict.get(id)
    for key in customers_dict:
        if key == listing.get_creatorID():
            seller = customers_dict.get(key)
            break

    # determine if user already liked this post
    customer = customers_dict.get(session_ID)  # current user
    customer_liked_posts = customer.get_liked_listings()
    user_liked_post = 'False'
    if listing.get_ID() in customer_liked_posts:
        user_liked_post = 'True'
    # get seller info
    # search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword=search_field.searchfield.data))
    except:
        pass
    # get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0
        # filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data, searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    product_title = listing.get_title()
    # estimated_amazon_price = get_amazon_estimated_price(product_title)
    estimated_ebay_price = get_ebay_estimated_price(product_title, listing.get_condition())
    estimated_amazon_price = 0
    # get username for navbar
    customer = customers_dict.get(session_ID)
    current_username = customer.get_username()

   
    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"


    return render_template('CustomerViewListing.html', listing=listing, seller=seller, current_sessionID=session_ID,
                           user_liked_post=user_liked_post, searchform=search_field,
                           customer_notifications=customer_notifications, filterform=filterform,
                           current_username=current_username, estimated_amazon_price=estimated_amazon_price,
                           estimated_ebay_price=estimated_ebay_price, )


#opstats - listing
@app.route('/reservelisting/<int:id>')
def reservelisting(id):
    dbmain = shelve.open('main.db','c')
    listings_dict = {}
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    listing = listings_dict.get(id)
    #opstats code
    try:
        if listing.get_status() == "available":
            Operatorstats.operatorstats_listings("total","minus")
            Operatorstats.operatorstats_listings("available","minus")
            Operatorstats.operatorstats_listings("reserved","plus")
        elif listing.get_status() == "disabled":
            Operatorstats.operatorstats_listings("total","minus")
            Operatorstats.operatorstats_listings("disabled","minus")
            Operatorstats.operatorstats_listings("reserved","plus")
        else:
            pass
    except:
        print("Error! Operator stats did not update.")
    listing.set_status('reserved')
    dbmain['Listings'] = listings_dict
    print("reserved listing")
    return redirect(url_for('confirmreservelisting', id=id))

#opstats - listing
@app.route('/unreservelisting/<int:id>')
def unreservelisting(id):
    dbmain = shelve.open('main.db','c')
    listings_dict = {}
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    listing = listings_dict.get(id)
    #opstats code
    try:
        Operatorstats.operatorstats_listings("available","plus")
        Operatorstats.operatorstats_listings("reserved","minus")
    except:
        print("Error! Operator stats did not update.")
    listing.set_status('available')
    dbmain['Listings'] = listings_dict
    return redirect(url_for('confirmunreservelisting', id=id))

@app.route('/reservelistingconfirmed/<int:id>')
def confirmreservelisting(id):
    global session_ID
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    listings_dict = {}
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    listing = listings_dict.get(id)

    #search func
    try:
        
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data)) #get the word from the search field
    except:
        pass
    
    #get notifs
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        customer_notifications = customer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if request.method == "POST" and filterform.validate():
            searchconditionlist = []
            get_searchquery(filterform.data,searchconditionlist)
            session['filters'] = searchconditionlist
            return redirect(url_for('filterresults'))
    except:
        pass
    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"
    return render_template('Customerconfirmreservelisting.html',current_sessionID = session_ID,searchform =search_field,customer_notifications = customer_notifications,filterform = filterform,listing = listing,current_username=current_username)

@app.route('/unreservelistingconfirmed/<int:id>')
def confirmunreservelisting(id):
    global session_ID
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    listings_dict = {}
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    listing = listings_dict.get(id)

    #search func
    try:
        
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data)) #get the word from the search field
    except:
        pass
    
    #get notifs
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        customer_notifications = customer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if request.method == "POST" and filterform.validate():
            searchconditionlist = []
            get_searchquery(filterform.data,searchconditionlist)
            session['filters'] = searchconditionlist
            return redirect(url_for('filterresults'))
    except:
        pass
    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"
    return render_template('Customerconfirmunreservelisting.html',current_sessionID = session_ID,searchform =search_field,customer_notifications = customer_notifications,filterform = filterform,listing = listing,current_username=current_username)

#opstatshere - listing
@app.route('/deleteListing/<int:id>/', methods = ['GET', 'POST'])
def deleteListing(id):
    global session_ID
    dbmain = shelve.open('main.db','c')
    listings_dict = {}
    customers_dict = {}
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    #get status of listing
    listing = listings_dict.get(id)
    try:
        if listing.get_status() == "available":
            Operatorstats.operatorstats_listings("total","minus")
            Operatorstats.operatorstats_listings("available","minus")
        elif listing.get_status() == "disabled":
            Operatorstats.operatorstats_listings("total","minus")
            Operatorstats.operatorstats_listings("disabled","minus")
        else:
            pass
    except:
        print("Error! Operator stats did not update.")
    del listings_dict[id]
    customer = customers_dict.get(session_ID)
    customer.remove_listings(id)
    print("listing has been removed")
    print(customer.get_listings())
    dbmain['Customers'] = customers_dict
    dbmain['Listings'] = listings_dict
    
    return redirect(url_for('Customerprofile', id=session_ID))

@app.route('/createReview/<int:id>', methods = ['GET', 'POST'])
def createReview(id):
    global session_ID
    review_form = ReviewForm(request.form)
    customers_dict = {}
    reviews_dict ={}
    notifications_dict = {}
    dbmain = shelve.open('main.db','c')

    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    try:
        if "Notifications" in dbmain:
            notifications_dict = dbmain["Notifications"] #sync local with db1
        else:
            dbmain['Notifications'] = notifications_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        dbmain = shelve.open('main.db','c')    
        Notifications.Notifications.count_ID = dbmain["NotificationsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Notifications count or count is at 0")
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Reviews" in dbmain:
            reviews_dict = dbmain["Reviews"] #sync local with db1
        else:
            dbmain['Reviews'] = reviews_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Reviews.Reviews.count_ID = dbmain["ReviewsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Review count or count is at 0")

    if request.method == 'POST' and review_form.validate():
        #get current user username
        current_customer = customers_dict.get(session_ID)
        current_customer_username = current_customer.get_username()
        

        #update reviews
        review = Reviews.Reviews(session_ID,current_customer_username,(float(review_form.rating.data)), review_form.review_text.data)
        reviews_dict[review.get_ID()] = review
        dbmain['Reviews'] = reviews_dict
        dbmain['ReviewsCount'] = Reviews.Reviews.count_ID
        print(f"Review has been added to db3\nReview ID:{review.get_ID()}\nReview creator_ID:{review.get_creator_ID()}\nReview rating:{review.get_rating()}\nReview comment:{review.get_comment()}")

        #update customer's reviews
        customer = customers_dict.get(id)#get opbject
        customer.add_reviews(review.get_ID())
        customer.set_rating(float(review.get_rating()))
        print(f"Customer reviews are {customer.get_reviews()}\n Customer current rating is {customer.get_rating()}")
        dbmain['Customers'] = customers_dict

        #notif
        #create notif obj
        notification = Notifications.Notifications(customer.get_id(),f"{current_customer_username} has left you a review!")
        notifications_dict[notification.get_ID()] = notification
        dbmain['Notifications'] = notifications_dict
        dbmain['NotificationsCount'] = Notifications.Notifications.count_ID

        #add notif to customer
        customer.add_notifications(notification.get_ID())
        dbmain['Customers'] = customers_dict
    
        #Email user
        Email.send_review_notifcation_gmail(customer.get_email(),customer.get_username(),current_customer_username,review.get_comment())
        return redirect(url_for('Customerprofile', id = id))
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    
    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"
    return render_template('CustomerReview.html',form=review_form, current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/createLikedListing/<int:id>', methods = ['GET', 'POST'])
def createLikedListing(id): #ID of listing
    global session_ID
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    listings_dict = {}
    notifications_dict = {}
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    #sync listing dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Listing.Listing.count_ID = dbmain["ListingsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Listing count or count is at 0")
    
    try:
        if "Notifications" in dbmain:
            notifications_dict = dbmain["Notifications"] #sync local with db1
        else:
            dbmain['Notifications'] = notifications_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        dbmain = shelve.open('main.db','c')    
        Notifications.Notifications.count_ID = dbmain["NotificationsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Notifications count or count is at 0")
    
    customer = customers_dict.get(session_ID) #get current user obj
    listing = listings_dict.get(id) #id that was entered

    #increment liked count of listing
    listing.add_likes() #add 
    dbmain['Listings'] = listings_dict
    print(f'Listing ID:{listing.get_ID()}, likes count is {listing.get_likes()}')

    #add liked post ID 
    customer.add_liked_listings(listing.get_ID())
    dbmain['Customers'] =  customers_dict
    print(f"Customer ID:{customer.get_id()} liked posts are {customer.get_liked_listings()}")
    
    #notif
    notification = Notifications.Notifications(listing.get_creatorID(),f"{customer.get_username()} liked your post:{listing.get_title()}")
    notifications_dict[notification.get_ID()] = notification
    dbmain['Notifications'] = notifications_dict
    dbmain['NotificationsCount'] = Notifications.Notifications.count_ID
    for key in notifications_dict:
        print(notifications_dict.get(key))

    #add notif to customer
    seller = customers_dict.get(listing.get_creatorID())
    seller.add_notifications(notification.get_ID())
    dbmain['Customers'] = customers_dict

    return redirect(url_for('viewListing', id = id))

#opstats, cancels delicvery
@app.route('/createUnlikedListing/<int:id>', methods = ['GET', 'POST'])
def createUnlikedListing(id):
    global session_ID
    dbmain = shelve.open('main.db','c')

    customers_dict = {} #local one
    listings_dict = {}
    deliveries_dict = {}

    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        Customer.Customer.count_id = dbmain["CustomerCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")


    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")

            # Sync deliveries
    try:
        if "delivery" in dbmain:
            deliveries_dict = dbmain["delivery"]
        else:
            dbmain['delivery'] = deliveries_dict
    except:
        print("Error in opening main.db")

    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Listing.Listing.count_ID = dbmain["ListingsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Listing count or count is at 0")
    
    customer = customers_dict.get(session_ID) #get current user obj
    listing = listings_dict.get(id) #id that was entered

    if not customer or not listing:
        print("Error: Customer or listing not found.")
        dbmain.close()
        return redirect(url_for('viewListing', id=id))

    if listing.get_likes() > 0:
        listing.minus_likes()
        print(f'Listing ID:{listing.get_ID()}, likes count is {listing.get_likes()}')

    customer.remove_liked_listings(id)

    for delivery in deliveries_dict.values():
        if delivery.get_item_title() == listing.get_title() and delivery.get_customer_id() == session_ID:
            #code for opstats
            if delivery.get_status() == "Pending":
                Operatorstats.operatorstats_feedbacks("Pending","minus")
                Operatorstats.operatorstats_feedbacks("Cancelled","plus")
            elif delivery.get_status() == "In Transit":
                Operatorstats.operatorstats_feedbacks("In Transit","minus")
                Operatorstats.operatorstats_feedbacks("Cancelled","plus")
            elif delivery.get_status() == "Delivered":
                Operatorstats.operatorstats_feedbacks("Delivered","minus")
                Operatorstats.operatorstats_feedbacks("Cancelled","plus")
            else:
                pass
            delivery.set_status("Cancelled")
            print(f'Delivery ID:{delivery.get_ID()} for Listing {listing.get_title()} is now Cancelled.')

    dbmain["Listings"] = listings_dict
    dbmain["Customers"] = customers_dict
    dbmain["delivery"] = deliveries_dict

    dbmain.close()

    return redirect(url_for('viewListing', id = id))

@app.route('/viewLikedListings/<int:id>', methods = ['GET', 'POST'])
def viewLikedListings(id): #retrieve current session_ID
    global session_ID
    dbmain = shelve.open('main.db','c')
    listings_dict = {}
    customers_dict = {}
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    customer = customers_dict.get(id) #current user
    customer_liked_listings = customer.get_liked_listings()
    listings_to_display = []
    for key in listings_dict:
        if key in customer_liked_listings:
            listing = listings_dict.get(key)
            listings_to_display.append(listing)
    #search function
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"
    return render_template('CustomerViewLikedListings.html', listings_to_display = listings_to_display, current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

#creates delivery object , opstats


@app.route('/delivery_status', methods=['GET', 'POST'])
def delivery_status():
    global session_ID
    dbmain = shelve.open('main.db', 'c')
    listings_dict = {}
    customers_dict = {}
    deliveries_dict = {}
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    form = DeliveryForm(request.form)

    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"]  # sync local with db2
        else:
            dbmain['Listings'] = listings_dict  # sync db2 with local (basically null)
    except:
        print("Error in opening main.db")

    # PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"]  # sync local with db1
        else:
            dbmain['Customers'] = customers_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        if "delivery" in dbmain:
            deliveries_dict = dbmain["delivery"]
        else:
            dbmain['delivery'] = deliveries_dict
    except:
        print("Error in opening main.db")
    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        Delivery.count_ID = dbmain["DeliveryCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Delivery count or count is at 0")

    customer = customers_dict.get(session_ID)

    liked_listings_ids = customer.get_liked_listings() if customer else []


    for listing_id in liked_listings_ids:
        listing = listings_dict.get(listing_id)
        if listing:
            # Check if the delivery for this item already exists
            existing_delivery = next((d for d in deliveries_dict.values() if
                                      d.get_item_title() == listing.get_title() and d.get_customer_id() == session_ID),
                                     None)

            if not existing_delivery:
                if customer:
                    expected_date = customer.get_date_joined()
                    # Add 5 days manually
                    month, day, year = map(int, expected_date.split('/'))


                    if year < 100:
                        year+= 2000
                    day += 5
                    # Handle month-end overflow
                    if month in [1, 3, 5, 7, 8, 10, 12] and day > 31:
                        day -= 31
                        month += 1
                    elif month in [4, 6, 9, 11] and day > 30:
                        day -= 30
                        month += 1
                    elif month == 2:  # Handle February (leap year check is skipped)
                        if day > 28:
                            day -= 28
                            month += 1

                    # Handle year-end overflow
                    if month > 12:
                        month = 1
                        year += 1

                    expected_date =f"{day:02d}/{month:02d}/{year%100:02d}"
                else:
                    expected_date = "TBD1"

                new_delivery = Delivery(
                    item_title=listing.get_title(),
                    status='Pending',  # Set status as 'Pending'
                    expected_date=expected_date,  # You can set an expected date or leave it as "TBD"
                    listing_id=session_ID,

                    address=listing.get_address()

                    #address =listing.get_deal_deliveryinfo()

                )


                deliveries_dict[new_delivery.get_ID()] = new_delivery

    dbmain["delivery"] = deliveries_dict
    dbmain['DeliveryCount'] = Delivery.count_ID

    #code for opstats
    try:
        Operatorstats.operatorstats_transactions("total","plus")
        Operatorstats.operatorstats_transactions("Pending","plus")
    except:
        pass

    deliveries_list = [
        delivery for delivery in deliveries_dict.values()
        if delivery.get_customer_id() == session_ID]
    print(deliveries_list)

    listings_to_display = []
    global_delivery_count = len(deliveries_dict)
    delivery = None

    # search function
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword=search_field.searchfield.data))
    except:
        pass

    # get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0
        # filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data, searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    
    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"

    return render_template('CustomerListingDelivery.html',delivery=delivery,form=form,deliveries_list=deliveries_list,customer=customer, listings_to_display=listings_to_display,
                           current_sessionID=session_ID, searchform=search_field,
                           customer_notifications=customer_notifications, filterform=filterform,current_username=current_username)

@app.route('/trackDelivery/<int:delivery_id>', methods=['GET', 'POST'])
def delivery_track(delivery_id):
    global session_ID
    dbmain = shelve.open('main.db', 'c')
    listings_dict = {}
    customers_dict = {}
    deliveries_dict = {}
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    form = DeliveryForm(request.form)

    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        Delivery.Delivery.count_id = dbmain["DeliveryCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Delivery count or count is at 0")
    if "DeliveryCount" in dbmain:
        Delivery.count_id = dbmain["DeliveryCount"]
    else:
        Delivery.count_id = 0
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"]  # sync local with db2
        else:
            dbmain['Listings'] = listings_dict  # sync db2 with local (basically null)
    except:
        print("Error in opening main.db")

    # PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"]  # sync local with db1
        else:
            dbmain['Customers'] = customers_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    try:
        if "delivery" in dbmain:
            deliveries_dict = dbmain["delivery"]
        else:
            dbmain['delivery'] = deliveries_dict
    except:
        print("Error in opening main.db")


    customer = customers_dict.get(session_ID, None)
    delivery = deliveries_dict.get(delivery_id)

    if delivery.get_customer_id() != session_ID:
        return render_template('error.html', message="Access denied."), 40
    delivery = deliveries_dict.get(delivery_id)
    if not delivery:
        return render_template('error.html', message="Delivery not found."), 404


    if customer:
        expected_date = customer.get_date_joined()
        # Add 5 days manually
        month, day, year = map(int, expected_date.split('/'))  # Assuming format is 'YYYY-MM-DD'

        if year < 100:
            year += 2000
        day += 5
        # Handle month-end overflow
        if month in [1, 3, 5, 7, 8, 10, 12] and day > 31:
            day -= 31
            month += 1
        elif month in [4, 6, 9, 11] and day > 30:
            day -= 30
            month += 1
        elif month == 2:  # Handle February (leap year check is skipped)
            if day > 28:
                day -= 28
                month += 1

        # Handle year-end overflow
        if month > 12:
            month = 1
            year += 1

        expected_date = f"{day:02d}/{month:02d}/{year%100:02d}"

    else:
        expected_date = "TBD"

    if request.method == "POST" and form.validate():
        # Update the delivery details
        delivery.set_address(form.deliveryinfo.data)
        delivery.set_expected_date(form.expected_date.data)

        # Save the updated delivery to the dict
        deliveries_dict[delivery.get_ID()] = delivery
        dbmain["delivery"] = deliveries_dict
        dbmain["DeliveryCount"] = Delivery.count_id
    dbmain.close()



    # get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0
        # filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data, searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"

    return render_template('Customerdeliverytrack.html', form=form, delivery=delivery,
                           customer=customer,deliveries_list=[delivery for delivery in deliveries_dict.values() if delivery.get_customer_id() == session_ID],
                           current_sessionID=session_ID, searchform=search_field,
                           customer_notifications=customer_notifications, filterform=filterform,current_username=current_username)


@app.route('/messages', methods=['GET', 'POST'])
def messages():
    global session_ID
    customers_dict = {}
    dbmain = shelve.open('main.db', 'c')
    # PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"]  # sync local with db1
        else:
            dbmain['Customers'] = customers_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    # add in for official project
    if session_ID == 0:
        return redirect(url_for('Customerhome'))
    db = shelve.open('messages.db', 'c')  # Open shelve database
    users_db = shelve.open('main.db')
    user = User(session_ID)
    recent_chats = user.get_recent_chats()
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    listings_dict = {}
    titles_dict = {}
    if "Listings" in dbmain:
        listings_dict = dbmain["Listings"]
    for key, i in listings_dict.items():
        titles_dict[key] = i.get_title()
    # get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0
        # search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword=search_field.searchfield.data))
    except:
        pass
    # filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data, searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"

    try:
        if request.method == 'POST':
            listingTag = False
            receiver_id = request.form.get('receiver_id', type=str)  # searchbar refers to this code when used, causes crash
            content = request.form.get('content', '').strip()  # Strip leading/trailing spaces
            customers_dict = users_db.get('Customers', {})
            reply_value = request.form.get('reply', type=str)
            reply_message_content = request.form.get('replymessagecontent', type=str)
            image_file = request.files.get('image')
            match = re.search(r'@(\d+)\s+([^\s]+)\b', content)
            reply_id = request.form.get('sender_ID', type=str)

            if match:
                listing_id = int(match.group(1))  # Extract the numeric ID
                listing_title = match.group(2)
                print(listing_title)
                if listing_id in titles_dict:
                    expected_title = titles_dict[listing_id]
                    num_words = len(expected_title.split())
                    adjusted_regex = r'@(\d+)\s+((?:\S+\s+){' + str(num_words - 1) + r'}\S+)\b'
                    match2 = re.search(adjusted_regex, content)
                    print(match2.group(2))
                    if match2.group(2) == expected_title:
                        listingTag = True
                        content = re.sub(adjusted_regex,f'<a href="/viewListing/{listing_id}">@{listing_id} {match2.group(2)}</a>', content, count=1)
            if not receiver_id.isdigit() or int(receiver_id) == 0 or int(receiver_id) not in customers_dict:
                return render_template(
                    'CustomerMessages.html',
                    error_message=f'User ID {receiver_id} is invalid or does not exist.',
                    show_error_modal=True,
                    current_sessionID=int(session_ID),
                    recent_chats=recent_chats,
                    selected_chat=None,
                    customer_notifications=customer_notifications,
                    searchform=search_field,
                    filterform=filterform
                )
            if int(receiver_id) == session_ID:
                return render_template(
                    'CustomerMessages.html',
                    error_message="You can't message yourself",
                    show_error_modal=True,
                    current_sessionID=int(session_ID),
                    recent_chats=recent_chats,
                    selected_chat=None,
                    customer_notifications=customer_notifications,
                    searchform=search_field,
                    filterform=filterform  # <- always put
                )
            receiver_id = int(receiver_id)  # Convert to integer after validation
            if image_file and content:
                message_type = "text+pic"
            elif image_file and content == '':
                message_type = "picture"
                content = ' '
            else:
                message_type = "text"
            if receiver_id and content == '':  # Only receiver_id is provided, no content
                return redirect(url_for('messages', receiver_id=receiver_id))  # Open the chat in the right column without changing recent chats order
            UUID = str(uuid.uuid4())
            filename = secure_filename(image_file.filename)
            if os.path.splitext(filename)[1] == ".png" or os.path.splitext(filename)[1] == ".jpg" or os.path.splitext(filename)[1] == ".jpeg":
                extension = os.path.splitext(filename)[1]
            else:
                message_type = "text"
                image_file = None
                extension = None
            reply_content = None
            sender_id = None
            if reply_value == "True" and reply_message_content != "" and reply_id != "":
                reply_content = reply_message_content
                sender_id = reply_id
            if customers_dict:
                receiver_object = customers_dict[receiver_id]
                receiver_name = receiver_object.get_username()
            if content:
                user.send_message(receiver_id, content, db, message_type, UUID, extension, None, reply_content, sender_id, receiver_name)
                if message_type == "text+pic" or message_type == "picture":
                    UPLOAD_FOLDER = "static/messagepics"
                    if not os.path.exists(UPLOAD_FOLDER):
                        os.makedirs(UPLOAD_FOLDER)
                    if image_file and image_file.filename != '':
                        unique_filename = UUID + os.path.splitext(filename)[1]
                        image_path = os.path.join(UPLOAD_FOLDER, unique_filename)  # Full path to save the file
                        image_file.save(image_path)  # Save the image
            if listingTag:
                global_listing = listings_dict[int(match.group(1))]
                message_type = "listingpic"
                content = (
                    f"<b>{global_listing.get_title()}</b>\n"
                    f"Seller: {global_listing.get_creator_username()}\n"
                    f"Status: {global_listing.get_status()}\n"
                    f"Condition: {global_listing.get_condition()}\n"
                    f"{global_listing.get_description()}\n"
                    f"<div><button onclick=\"window.location.href='/viewListing/{int(match.group(1))}'\" id=\"viewlistingbutton\">View listing</button></div>"
                )
                UUID = str(uuid.uuid4())
                folder_path = "static/listingpics"  # Path where images are stored
                filename_prefix = f"listing{int(match.group(1))}"  # The expected filename pattern
                for file in os.listdir(folder_path):
                    if file.startswith(filename_prefix):  # Check if the file starts with "listing{ID}"
                        file_extension = os.path.splitext(file)[1]  # Extract extension
                        extension = file_extension
                        break  # Stop after finding the first match
                    else:
                        extension = '.jpg'
                user.send_message(receiver_id, content, db, message_type, UUID, extension, int(match.group(1)), None, None, receiver_name)
            return redirect(url_for('messages', receiver_id=receiver_id))  # Reload to show the new message

        received_messages = user.get_received_messages(db)
        sent_messages = user.get_sent_messages(db)
        selected_chat = None
        customers_dict = users_db.get('Customers', {})
        if 'receiver_id' in request.args:
            receiver_id = request.args.get('receiver_id', type=int)
            sender_id = session_ID
            receiver_name = None
            if customers_dict:
                receiver_object = customers_dict[receiver_id]
                receiver_name = receiver_object.get_username()

            message = [
                {
                    "type": "sent" if message.sender_id == session_ID else "received",
                    "content": message.content,
                    "timestamp": message.timestamp.strftime("%Y-%m-%d %H:%M"),
                    "receiver_id": message.receiver_id,
                    "sender_id": message.sender_id,
                    "message_id": message.message_id,
                    "status": message.status,
                    "type2": message.type,
                    "extension": message.extension,
                    "listing_id": str(message.listing_id),
                    "reply_content": message.reply_value,
                    "sender_reply_id": message.replier_id,
                    "receiver_name": receiver_name
                }
                for message in db.get("Messages", [])
                if (message.sender_id == session_ID and message.receiver_id == receiver_id) or
                   (message.receiver_id == session_ID and message.sender_id == receiver_id)
            ]

            selected_chat = {
                'receiver_id': receiver_id,
                'messages': message,
                'sender_id': sender_id,
                'receiver_name': receiver_name
            }
        return render_template(
            'CustomerMessages.html',
            received_messages=received_messages,
            sent_messages=sent_messages,
            current_sessionID=session_ID,
            recent_chats=recent_chats,
            selected_chat=selected_chat,
            searchform=search_field,
            customer_notifications=customer_notifications,
            show_error_modal=False,
            filterform=filterform,
            listings_dict=listings_dict,
            titles_dict=titles_dict
        )
    finally:
        db.close()
@app.route('/delete_chat', methods=['POST'])
def delete_chat():
    db = shelve.open('recentChat.db', 'c')
    user = User(session_ID)
    recent_chats = user.get_recent_chats()
    data = request.get_json()  # Get the JSON data
    receiver_id = data.get('receiver_id')
    # Process the receiver_id as needed
    print(recent_chats)
    print(f"Received receiver_id: {receiver_id}")
    updated_chats = []
    for chat in recent_chats:
        if not chat['receiver_id'] == int(receiver_id):
            updated_chats.append(chat)
    print(updated_chats)
    db[str(session_ID)] = updated_chats
    return jsonify({'message': f'Receiver ID {receiver_id} processed successfully.'})

@app.route('/delete_message', methods=['POST'])
def delete_message():

    data = request.get_json()  # Get the JSON data from the request
    message_id = data.get('message_id')  # Get the message ID

    db = shelve.open('messages.db', 'c')
    messages2 = db.get('Messages', [])
    message_to_delete = None

    for message in messages2:
        if message.message_id == message_id:
            message_to_delete = message
            break

    if not message_to_delete:
        return jsonify({'success': False, 'error': 'Message not found'}), 404

    # Delete the message from the database
    # messages2.remove(message_to_delete)
    message_to_delete.status = "deleted"
    message_to_delete.content = None #privacy concerns
    db['Messages'] = messages2

    db.close()


    # to do: ask teacher if it counts as "deletion" of database to do the above, must there be real time messages(current code only shows message upon refresh),
    # cannot enter own ID for new chat, option to edit/reply to messages
    # dropdown menu: option to delete chat/hyperlink to user's profile/block profile,
    # option to send pictures in chat, message delivered/read/notifications(red number icon),
    # message previews, make date appear like whatsapp

@app.route('/edit_message', methods=['GET', 'POST'])
def edit_message():
    data = request.get_json()  # Safely parse the JSON data
    messageId = data.get('messageId')
    newMessage = data.get('message')

    print(f"Received messageId: {messageId}, new message: {newMessage}")
    db = shelve.open('messages.db', 'c')
    try:
        messages_list = db.get("Messages", [])  # Retrieve the list of messages
        message_found = False
        for message in messages_list:
            if message.message_id == messageId:  # Match the message ID
                message.content = newMessage
                message.status = "edited"
                message_found = True
                break

        if not message_found:
            return jsonify({"error": "Message not found"}), 404

        # Save the updated messages list back to the database
        db["Messages"] = messages_list

        return jsonify({
            "message": "Message updated successfully",
            "messageId": messageId,
            "newContent": newMessage
        }), 200
    finally:
        db.close()

#listingalgo
@app.route('/searchresults/<keyword>', methods=['GET', 'POST'])
def searchresults(keyword):
    global session_ID
    dbmain = shelve.open('main.db','c')
    customers_dict = {}
    listings_dict = {}
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db") 

    show_listings = []
    for key in listings_dict:
        listing = listings_dict.get(key)
        
        Search.search_keyword(listing,keyword,show_listings)#check if it fulfills the condition
    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    
    #get username for navbar
    if session_ID != 0:
        current_customer = customers_dict.get(session_ID)
        current_username = current_customer.get_username()
    else:
        current_username = "nil"
    return render_template("Customersearchresults.html",current_sessionID = session_ID,searchform =search_field,listings_list = show_listings,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)
    
@app.route('/category1', methods=['GET', 'POST'])
def category1():
    global session_ID
    customers_dict = {}
    listings_dict = {}
    dbmain = shelve.open('main.db','c')
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #search
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #get prods from cat1
    listings_to_display = []
    for key in listings_dict:
        listing = listings_dict.get(key)
        if listing.get_category() == 'Category 1' and Search.check_listing(listing):

            listings_to_display.append(listing)
    
    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"

    return render_template('CustomerCategory1.html', listings_list = listings_to_display, current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/category2',methods = ['GET','POST'])
def category2():
    global session_ID
    customers_dict = {}
    listings_dict = {}
    dbmain = shelve.open('main.db','c')

    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #search
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #get prods from cat1
    listings_to_display = []
    for key in listings_dict:
        listing = listings_dict.get(key)
        if listing.get_category() == 'Category 2' and Search.check_listing(listing) :
            listings_to_display.append(listing)

    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"

    return render_template('CustomerCategory2.html', listings_list = listings_to_display, current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/category3',methods = ['GET','POST'])
def category3():
    global session_ID
    customers_dict = {}
    listings_dict = {}
    dbmain = shelve.open('main.db','c')
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #search
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #get prods from cat1
    listings_to_display = []
    for key in listings_dict:
        listing = listings_dict.get(key)
        if listing.get_category() == 'Category 3' and Search.check_listing(listing) :
            listings_to_display.append(listing)

    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"
    return render_template('CustomerCategory3.html', listings_list = listings_to_display, current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/category4',methods = ['GET','POST'])
def category4():
    global session_ID
    customers_dict = {}
    listings_dict = {}
    dbmain = shelve.open('main.db','c')
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #search
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #get prods from cat1
    listings_to_display = []
    for key in listings_dict:
        listing = listings_dict.get(key)
        if listing.get_category() == 'Category 4' and Search.check_listing(listing) :
            listings_to_display.append(listing)
    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"
    return render_template('CustomerCategory4.html', listings_list = listings_to_display, current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/category5',methods = ['GET','POST'])
def category5():
    global session_ID
    customers_dict = {}
    listings_dict = {}
    dbmain = shelve.open('main.db','c')

    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #search
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #get prods from cat1
    listings_to_display = []
    for key in listings_dict:
        listing = listings_dict.get(key)
        if listing.get_category() == 'Category 5' and Search.check_listing(listing) :
            listings_to_display.append(listing)
    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"
    return render_template('CustomerCategory5.html', listings_list = listings_to_display, current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

@app.route('/filterresults/',methods = ['GET','POST'])
def filterresults():
    global session_ID
    customers_dict = {}
    dbmain = shelve.open('main.db','c')
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening customer.db")
    #search
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #func
    outputlistID = []
    listings_to_display = []
    get_matchinglistingID(session['filters'],outputlistID)
    outputlistID = deduper(outputlistID)
    ID_to_obj(outputlistID,listings_to_display)
    
    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"
    
    return render_template('Customerfilterresults.html',listings_list = listings_to_display, current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,current_username=current_username)

#opstats - feedback
@app.route('/feedback', methods = ['GET', 'POST'])
def feedback():
    global session_ID
    customers_dict = {}
    dbmain  = shelve.open('main.db','c')
    feedbacks_dict = {}
    search_field = SearchBar(request.form)
    feedback_form = FeedbackForm(request.form)
    filterform = FilterForm(request.form)
    reply_feedback_form = ReplyFeedback(request.form)
    reply = reply_feedback_form.reply.data if reply_feedback_form.validate() else None
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        if "Feedback" in dbmain:
            feedbacks_dict = dbmain["Feedback"] #sync local with db1
        else:
            dbmain['Feedback'] = feedbacks_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Feedback.Feedback.count_ID = dbmain["FeedbackCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Feedback count or count is at 0")

    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0 

    if request.method == 'POST' and feedback_form.validate():
        #add the feedback into feedback db
        reply = reply_feedback_form.reply.data if reply_feedback_form.validate() else None
        feedback = Feedback.Feedback(feedback_form.rating.data,feedback_form.feedback.data,reply)
        feedbacks_dict[feedback.get_ID()] = feedback #store obj in dict
        dbmain['Feedback'] = feedbacks_dict
        dbmain['FeedbackCount'] = Feedback.Feedback.count_ID
        try:
            Operatorstats.operatorstats_feedbacks("total","plus")
            Operatorstats.operatorstats_feedbacks("unreplied","plus")
        except:
            pass
        #add feedback ID to customer
        customer = customers_dict.get(session_ID)
        if customer:
            customer.add_feedback(feedback.get_ID())
            dbmain['Customers'] = customers_dict


        return redirect(url_for('Customerhome'))


    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"
    
    return render_template('CustomerFeedback.html',searchform =search_field,
            customer_notifications=customer_notifications,current_sessionID=int(session_ID),form=feedback_form,filterform=filterform,current_username=current_username)

@app.route('/notifications/<int:id>')
def viewnotifications(id): #id is current_sessionID
    global session_ID
    dbmain = shelve.open('main.db','c')
    notifications_dict = {}
    customers_dict = {}
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        if "Notifications" in dbmain:
            notifications_dict = dbmain["Notifications"] #sync local with db1
        else:
            dbmain['Notifications'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    
    try:
        dbmain = shelve.open('main.db','c')    
        Notifications.Notifications.count_ID = dbmain["NotificationsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main notifs count or count is at 0")

    #get own notifs
    customer = customers_dict.get(id)
    notifs_list = customer.get_notifications()
    notifs_to_display = []
    for key in notifications_dict:
        if key in notifs_list:
            notification = notifications_dict.get(key)
            notifs_to_display.append(notification)
    #set unread notifs to 0
    customer.clear_unread_notifications()
    dbmain['Customers'] = customers_dict
    dbmain.close()

    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass

    # get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0

    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"
    return render_template("CustomerViewNotifications.html",current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,notifications_list = notifs_to_display,filterform=filterform,current_username=current_username)
    
@app.route('/profilefeedback/<int:id>', methods = ['GET', 'POST'])
def Customerprofilefeedback(id):#id not needed for now
    dbmain = shelve.open('main.db', 'c')
    customers_dict = {}  # local one
    global session_ID
    feedbacks_dict = {}
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    updatefeedbackform = UpdateFeedback(request.form)
    reply_feedback_form = ReplyFeedback(request.form)
    try:
        if "Feedback" in dbmain:
            feedbacks_dict = dbmain["Feedback"]  # sync local with db1
        else:
            dbmain['Feedback'] = feedbacks_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        # make sure local and db1 are the same state
    # PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"]  # sync local with db1
        else:
            dbmain['Customers'] = customers_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    #get current customer
    customer = customers_dict.get(session_ID)
    customer_feedbacks_list = customer.get_feedbacks()
    feedbacks_list = []
    for key in feedbacks_dict:
        if key in customer_feedbacks_list: #ensure is own customer
            feedback = feedbacks_dict.get(key)
            feedbacks_list.append(feedback)
            if not hasattr(feedback, "_Feedback__category"):
                feedback._Feedback__category = "General"

    numberfeedbacks = len(feedbacks_list)
    
    
    # search func
    try:

        if request.method == 'POST' and search_field.validate():
            return redirect(
                url_for('searchresults', keyword=search_field.searchfield.data))  # get the word from the search field
    except:
        pass

    # get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0

    # filter
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    
    print(feedbacks_list)
    
    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"
    return render_template('Customerprofile_feedback.html',number_of_feedbacks = numberfeedbacks,list_feedback = feedbacks_list, current_sessionID=session_ID,searchform=search_field
                           ,customer_notifications=customer_notifications,reply_feedback_form=reply_feedback_form, filterform=filterform, customer=customer,updatefeedbackform = updatefeedbackform,current_username=current_username)

@app.route('/update_feedback/<int:feedback_id>', methods=['POST', 'GET'])
def update_feedback(feedback_id):
    feedbacks_dict = {}
    dbmain = shelve.open('main.db', 'c')
    customers_dict = {}  # local one
    global session_ID
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    
    # Get the feedback dictionary
    try:
        if "Feedback" in dbmain:
            feedbacks_dict = dbmain["Feedback"]  # sync local with db1
        else:
            dbmain['Feedback'] = feedbacks_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        Feedback.Feedback.count_ID = dbmain["FeedbackCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Feedback count or count is at 0")
    # PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"]  # sync local with db1
        else:
            dbmain['Customers'] = customers_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")


    # Check if the feedback ID exists
    if feedback_id not in feedbacks_dict:
        return "Feedback not found", 404
    feedback = feedbacks_dict.get(feedback_id)
    update_feedback_form = UpdateFeedback(request.form,rating = feedback.get_rating(),feedback = feedback.get_remark())
    
    print(f"Feedback id:{feedback_id}")  # Retrieve the feedback obj
    if request.method == 'POST' and update_feedback_form.validate():
        # Update feedback details
        feedback.set_rating(update_feedback_form.rating.data)
        feedback.set_remark(update_feedback_form.feedback.data)
        feedback.set_category(update_feedback_form.category.data)
        dbmain['Feedback'] = feedbacks_dict  # Update the database

        return redirect(url_for('Customerprofilefeedback', id = session_ID))  # Redirect to the feedback dashboard
    
    feedbacks_list = []
    #idk what this part below does anyways
    for key in feedbacks_dict:
        feedback = feedbacks_dict.get(key)
        feedbacks_list.append(feedback) #get all feedback obj
    # make sure local and db1 are the same state
    
    # search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(
                url_for('searchresults', keyword=search_field.searchfield.data))  # get the word from the search field
    except:
        pass
    # get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass

    
    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"
    return render_template("CustomerUpdatefeedback.html",current_sessionID = session_ID,feedback=feedback,feedbacks_list=feedbacks_list,update_feedback_form=update_feedback_form,searchform =search_field,customer_notifications = customer_notifications,filterform = filterform,current_username=current_username)

#opstats - feedback
@app.route('/delete_feedback/<int:feedback_id>', methods=['POST','GET'])
def delete_feedback(feedback_id):
    global session_ID
    print('reached delete feedback')
    dbmain = shelve.open('main.db', 'c')
    feedbacks_dict = dbmain.get('Feedback', {})  # Get the feedback dictionary
    customers_dict = {}
    try:
        if "Feedback" in dbmain:
            feedbacks_dict = dbmain["Feedback"]  # sync local with db1
        else:
            dbmain['Feedback'] = feedbacks_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    # PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"]  # sync local with db1
        else:
            dbmain['Customers'] = customers_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        Feedback.Feedback.count_ID = dbmain["FeedbackCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Feedback count or count is at 0")
    feedback = feedbacks_dict.get(feedback_id)
    try:
        if feedback.get_reply() == None:
            Operatorstats.operatorstats_feedbacks("total","minus")
            Operatorstats.operatorstats_feedbacks("unreplied","minus")
        else:
            Operatorstats.operatorstats_feedbacks("total","minus")
            Operatorstats.operatorstats_feedbacks("replied","minus")
    except:
        pass
    # Check if the feedback ID exists
    if feedback_id in feedbacks_dict:
        del feedbacks_dict[feedback_id]  # Remove the feedback
        dbmain['Feedback'] = feedbacks_dict# Save changes to the database

    customer = customers_dict.get(session_ID)
    customer.remove_feedback(feedback_id)
    dbmain['Customers'] = customers_dict

    dbmain.close()
    
    return redirect(url_for('Customerprofilefeedback',id=session_ID))

#opstats - feedback
@app.route('/reply_feedback/<int:feedback_id>', methods=['POST', 'GET'])
def reply_feedback(feedback_id):
    feedbacks_dict = {}

    dbmain = shelve.open('main.db', 'c')
    customers_dict = {}  # local one
    global session_ID
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    feedback = feedbacks_dict.get(feedback_id)
    reply = request.form.get('reply')
    reply_feedback_form = ReplyFeedback(request.form)
    # Get the feedback dictionary


    try:
        if "Feedback" in dbmain:
            feedbacks_dict = dbmain["Feedback"]  # sync local with db1
        else:
            dbmain['Feedback'] = feedbacks_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        Feedback.Feedback.count_ID = dbmain["FeedbackCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Feedback count or count is at 0")
    # PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"]  # sync local with db1
        else:
            dbmain['Customers'] = customers_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    # Check if the feedback ID exists
    if feedback_id not in feedbacks_dict:
        return "Feedback not found", 404
    
    try:
        Operatorstats.operatorstats_feedbacks("unreplied","minus")
        Operatorstats.operatorstats_feedbacks("replied","plus")
    except:
        pass
    print(f"Feedback id:{feedback_id}")  # Retrieve the feedback obj
    if request.method == 'POST' and reply_feedback_form.validate():
        # Update feedback details
        feedback = feedbacks_dict.get(feedback_id)
        feedback.set_reply(reply)
        dbmain['Feedback'] = feedbacks_dict  # Update the database

        return redirect(url_for('dashboardfeedbacks'))  # Redirect to the feedback dashboard

    feedbacks_list = []
    # idk what this part below does anyways
    for key in feedbacks_dict:
        feedback = feedbacks_dict.get(key)
        feedbacks_list.append(feedback)  # get all feedback obj
    # make sure local and db1 are the same state

    # search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(
                url_for('searchresults', keyword=search_field.searchfield.data))  # get the word from the search field
    except:
        pass
    # get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0
    # filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data, searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    




    return render_template("Operatordashboard_feedback_reply.html", current_sessionID=session_ID, feedback=feedback,
                           feedbacks_list=feedbacks_list, reply_feedback_form=reply_feedback_form,
                           searchform=search_field, customer_notifications=customer_notifications,
                           filterform=filterform)

#report user
@app.route('/report/user/<int:id>', methods=['POST', 'GET'])
def report_user(id):
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    global session_ID
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")


    #logic to handle ID = 0
    if id != 0 :
        customer = customers_dict.get(id)
        customer_username = customer.get_username()
        reportform = ReportForm(request.form,affectedusername = customer_username)
    elif id == 0:
        reportform = ReportForm(request.form)
    
    if request.method == 'POST' and reportform.validate():
        
        #check if useranem exists
        flag = False
        for key in customers_dict:
            if (reportform.affectedusername.data) == ((customers_dict.get(key)).get_username()):
                print("username true")
                session['affectedid']  = ((customers_dict.get(key)).get_id())
                flag = True
                break
            else:
                flag = False
            
        if flag == True:
            session['report_category'] = reportform.category.data
            session['report_text'] = reportform.report_text.data
            return(redirect(url_for('confirmreportuser',id=int(session['affectedid']))))
        else:
            return(redirect(url_for('cinvaliduser')))
    else:
        pass


    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data)) #get the word from the search field
    except:
        pass
    #get notifs
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        customer_notifications = customer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    #get username for navbar
    customer = customers_dict.get(session_ID)
    current_username = customer.get_username()
    return render_template('Customerreport.html', current_sessionID = session_ID,searchform =search_field,customer_notifications = customer_notifications,filterform = filterform,reportform = reportform,customerID= id,current_username=current_username)

@app.route('/confirmreport/user/<int:id>', methods=['POST', 'GET'])
def confirmreportuser(id):
    global session_ID
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    report_form = ReportForm(request.form)
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    

    customer = customers_dict.get(id)

    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data))
    except:
        pass
    #get notifs
    if session_ID != 0:
        owncustomer = customers_dict.get(session_ID)
        customer_notifications = owncustomer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    #get username for navbar
    customer = customers_dict.get(session_ID)
    current_username = customer.get_username()
    return render_template('Customerconfirmreport.html',customer = customer,current_sessionID = session_ID,form=report_form,searchform =search_field,customer_notifications = customer_notifications,filterform=filterform,customerid = id,current_username=current_username)

@app.route('/reportuser/<int:id>', methods=['GET', 'POST'])
def report(id):
    global session_ID
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    reports_dict = {}
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    #sync report dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Reports" in dbmain:
            reports_dict = dbmain["Reports"] #sync local with db2
        else:
            dbmain['Reports'] = reports_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Report.Report.count_ID = dbmain["ReportsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Report count or count is at 0")
    customer = customers_dict.get(id)
    report = Report.Report(session_ID,id,customer.get_username(),session['report_category'],session['report_text'])
    reports_dict[report.get_ID()] = report #store obj in dict
    dbmain['Reports'] = reports_dict
    dbmain['ReportsCount'] = Report.Report.count_ID

    #store report in offender's report_listings
    customer = customers_dict.get(id)
    customer.add_reports(report.get_ID())
    dbmain['Customers'] = customers_dict

    try:
        Operatorstats.operatorstats_reports("total","plus")
    except:
        pass

    dbmain.close()

    return redirect(url_for('Customerprofile', id=id))

@app.route('/c_invaliduser',methods=['GET','POST'])
def cinvaliduser():
    dbmain = shelve.open('main.db','c')
    customers_dict = {} #local one
    global session_ID
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    #search func
    try:
        
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data)) #get the word from the search field
    except:
        pass
    
    #get notifs
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        customer_notifications = customer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if request.method == "POST" and filterform.validate():
            searchconditionlist = []
            get_searchquery(filterform.data,searchconditionlist)
            session['filters'] = searchconditionlist
            return redirect(url_for('filterresults'))
    except:
        pass
    #get username for navbar
    customer = customers_dict.get(session_ID)
    current_username = customer.get_username()
    return render_template('Customerinvaliduser.html', current_sessionID = session_ID,searchform =search_field,customer_notifications = customer_notifications,filterform = filterform,current_username=current_username)


@app.route('/addToCart/<int:id>/', methods = ['GET', 'POST'])
def addToCart(id):
    global session_ID
    dbmain = shelve.open('main.db', 'c')
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    listings_dict = {}
    customers_dict = {}
    notifications_dict = {}
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    #sync listing dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')
        Listing.Listing.count_ID = dbmain["ListingsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Listing count or count is at 0")

    try:
        if "Notifications" in dbmain:
            notifications_dict = dbmain["Notifications"] #sync local with db1
        else:
            dbmain['Notifications'] = notifications_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        dbmain = shelve.open('main.db','c')
        Notifications.Notifications.count_ID = dbmain["NotificationsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Notifications count or count is at 0")


    customer = customers_dict.get(session_ID)
    listing = listings_dict.get(id)

    if request.method == 'POST':
        deal_method = request.form.get("deal_method")
        address = request.form.get("address")  # Always take the address

        if not deal_method:
            dbmain.close()
            return render_template("CustomerDealMethod.html", listing=listing)

        listing.set_deal_method(deal_method)
        listing.set_address(address)
        customer.add_cart_listing(listing.get_ID())

        dbmain["Listings"] = listings_dict
        dbmain["Customers"] = customers_dict
        dbmain.close()

        return redirect(url_for('viewCart'))

    dbmain.close()
    #search func
    try:
        
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data)) #get the word from the search field
    except:
        pass
    
    #get notifs
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        customer_notifications = customer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if request.method == "POST" and filterform.validate():
            searchconditionlist = []
            get_searchquery(filterform.data,searchconditionlist)
            session['filters'] = searchconditionlist
            return redirect(url_for('filterresults'))
    except:
        pass
    return render_template('CustomerDealMethod.html',current_sessionID = session_ID,searchform =search_field,customer_notifications=customer_notifications,filterform=filterform,listing = listing)

# - no form validation so POST request goes to other handler meaning kaboom
@app.route('/viewCart/<int:id>/', methods = ['GET', 'POST'])
def viewCart(id):
    global session_ID
    dbmain = shelve.open('main.db', 'c')
    listings_dict = {}
    customers_dict = {}
    deliveries_dict = {}
    notifications_dict = {}
    search_field = SearchBar(request.form)
    filterform = FilterForm(request.form)
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    #sync listing dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')
        Listing.Listing.count_ID = dbmain["ListingsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Listing count or count is at 0")

    try:
        if "Notifications" in dbmain:
            notifications_dict = dbmain["Notifications"] #sync local with db1
        else:
            dbmain['Notifications'] = notifications_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    try:
        dbmain = shelve.open('main.db','c')
        Notifications.Notifications.count_ID = dbmain["NotificationsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Notifications count or count is at 0")

    try:
        if "delivery" in dbmain:
            deliveries_dict = dbmain["delivery"]
        else:
            dbmain['delivery'] = deliveries_dict
    except:
        print("Error in opening deliery in main.db")
    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        Delivery.count_ID = dbmain["DeliveryCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Delivery count or count is at 0")

    # Fetch the cart listings for the customer
    customer = customers_dict.get(id)
    selectedcustomer_cart_list = customer.get_cart_listings()
    # Passing data to template
    #search func
    try:
        if request.method == 'POST' and search_field.validate():
            return redirect(url_for('searchresults', keyword = search_field.searchfield.data)) #get the word from the search field
    except:
        pass
    #get notifs
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        customer_notifications = customer.get_unread_notifications()
    elif session_ID == 0:
        customer_notifications = 0  
    #filter
    try:
        if filterdict(filterform.data) == True:
            if request.method == "POST" and filterform.validate():
                searchconditionlist = []
                get_searchquery(filterform.data,searchconditionlist)
                session['filters'] = searchconditionlist
                return redirect(url_for('filterresults'))
        else:
            pass
    except:
        pass
    #get username for navbar
    if session_ID != 0:
        customer = customers_dict.get(session_ID)
        current_username = customer.get_username()
    else:
        current_username = "nil"
    # Prepare the listings to display
    listings_to_display = []
    for listing_id in selectedcustomer_cart_list:
        currentlisting = listings_dict.get(int(listing_id))
        if currentlisting:
            # Filter listings with deal method "delivery"
            if currentlisting.get_deal_method().lower() == "delivery":
                expected_date = customer.get_date_joined()
                month, day, year = map(int, expected_date.split('/'))

                if year < 100:
                    year += 2000
                day += 5  # Add 5 days

                # Handle month-end and year-end overflow
                if month in [1, 3, 5, 7, 8, 10, 12] and day > 31:
                    day -= 31
                    month += 1
                elif month in [4, 6, 9, 11] and day > 30:
                    day -= 30
                    month += 1
                elif month == 2 and day > 28:
                    day -= 28
                    month += 1
                if month > 12:
                    month = 1
                    year += 1

                selectedexpecteddate = f"{day:02d}/{month:02d}/{year % 100:02d}"
                selectedaddress = currentlisting.get_address() or "N/A"
                selectedtitle = currentlisting.get_title()

                # Create the delivery object
                deliveryobj = Delivery.Delivery(
                    selectedID=currentlisting.get_ID(),
                    item_title=selectedtitle,
                    status="Pending",
                    expected_date=selectedexpecteddate,
                    address=selectedaddress
                )

                # Add the delivery to the deliveries dictionary
                deliveries_dict[deliveryobj.get_ID()] = deliveryobj
                dbmain["delivery"] = deliveries_dict

                # Log for debugging
                print(f"Delivery Info: ID={deliveryobj.get_ID()}, Title={deliveryobj.get_item_title()}, Status={deliveryobj.get_status()}, Expected Date={deliveryobj.get_expected_date()}, Address={deliveryobj.get_address()}")

                # Add to listings to display for Jinja template
                listings_to_display.append(currentlisting)

    dbmain.close()

    
    return render_template('CustomerviewCart.html', listings_to_display=listings_to_display, delivery_cost=5,current_sessionID = session_ID,searchform =search_field,customer_notifications = customer_notifications,filterform = filterform,current_username=current_username)

@app.route('/removeFromCart/<int:item_id>', methods=['POST'])
def removeFromCart(item_id):
    global session_ID
    dbmain = shelve.open('main.db', 'c')

    customers_dict = {}
    listings_dict = {}

    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    #sync listing dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')
        Listing.Listing.count_ID = dbmain["ListingsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Listing count or count is at 0")

    customer = customers_dict.get(session_ID)
    if not customer:
        dbmain.close()
        return redirect(url_for('viewCart'))

    # Remove the item from the customer's cart if it exists.
    # (Assuming cart_listings is a list of listing IDs.)
    cart = customer.get_cart_listings()
    if item_id in cart:
        cart.remove(item_id)
        print(f'Item ID:{item_id} removed from the cart.')
    listing = listings_dict.get(item_id)
    listing.set_deal_method(None)
    listing.set_address(None)

    # Save updates back to the database
    dbmain["Customers"] = customers_dict
    dbmain["Listings"] = listings_dict
    dbmain.close()

    return redirect(url_for('viewCart'))

@app.route('/loginoperator', methods=['GET', 'POST'])
def loginoperator():
    global session_ID
    search_field = SearchBar(request.form)
    operator_login_form = OperatorLoginForm(request.form)
    OTP = False #set it to false if you dont want to use the OTP feature
    if request.method == 'POST' and operator_login_form.validate():
        session['Email'] = operator_login_form.email.data #stores the data
        if operator_login_form.operator_username.data == "sysadmin1":
            if operator_login_form.password.data == "sysadmin1":
                if OTP == True:
                    OTP = ' '.join([str(random.randint(0,999999)).zfill(6)])
                    session['OTP'] = OTP
                    #print(f"Session OTP is {session['OTP']}")
                    Email.send_message_operator_OTP(operator_login_form.email.data,OTP)
                    return redirect(url_for('verifyoperator', email = operator_login_form.email.data))
                elif OTP == False:
                    return(redirect(url_for('dashboardusers')))
                
    return render_template('Operatorlogin.html',searchform =search_field, form=operator_login_form)

@app.route('/verifyoperator/<email>',methods=['GET', 'POST'])
def verifyoperator(email):
    search_field = SearchBar(request.form)
    operator_OTP = OperatorLoginVerifyForm(request.form)

    if request.method == 'POST' and operator_OTP.validate():#submit action

        print(operator_OTP.OTP.data)
        if session['OTP'] == operator_OTP.OTP.data:
            return(redirect(url_for('dashboardusers')))
        else:#IF first time fail then
            OTP = ' '.join([str(random.randint(0,999999)).zfill(6)])
            session['OTP'] = OTP
            print(f"Session OTP is {session['OTP']}")
            Email.send_message_operator_OTP(email,OTP)
            return redirect(url_for('verifyoperator', email = email))
    
    
    return render_template('OperatorLoginVerify.html', searchform =search_field,form = operator_OTP)

@app.route('/operatorcontrolcenter')
def operatorcontrolcenter():
    dbmain = shelve.open('main.db', 'c')
    operatorstats_dict = {}
    try:
        if "Operatorstats" in dbmain:
            operatorstats_dict = dbmain["Operatorstats"]  # sync local with db1
        else:
            dbmain['Operatorstats'] = operatorstats_dict # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    operator_stats = operatorstats_dict.get(1)
    userinfo=[operator_stats.get_users_count(),operator_stats.get_users_active_count(),operator_stats.get_users_suspended_count(),operator_stats.get_users_terminated_count()]
    listinginfo=[operator_stats.get_listings_count(),operator_stats.get_listings_available_count(),operator_stats.get_listings_disabled_count()]
    feedbackinfo = [operator_stats.get_feedback_count(),operator_stats.get_feedback_replied_count(),operator_stats.get_feedback_unreplied_count()]
    transactioninfo = [operator_stats.get_transactions_count(),operator_stats.get_transactions_Pending_count(),operator_stats.get_transactions_In_Transit_count(),operator_stats.get_transactions_Delivered_count(),operator_stats.get_transactions_Cancelled_count()]
    reportinfo = [operator_stats.get_reports_count()]
    opactioninfo = [operator_stats.get_opactions_count()]
    return render_template('OperatorControlCenter.html',userinfo = userinfo,listinginfo=listinginfo,feedbackinfo = feedbackinfo,transactioninfo = transactioninfo,reportinfo = reportinfo,opactioninfo = opactioninfo)

#dashboard users
@app.route('/dashboard/users',methods=['GET', 'POST'])
def dashboardusers():
    user_search_field = SearchUserField(request.form)
    user_status_field = SearchUserStatus(request.form)
    dbmain = shelve.open('main.db','c')
    customers_dict = {}
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")
    customers_list = []
    for key in customers_dict:
        customer = customers_dict.get(key)
        customers_list.append(customer)
    try:

        if request.method == 'POST' and user_search_field.validate():
            return redirect(url_for('dashboarduser_usernamesearch',keyword = user_search_field.searchfield.data))
    except:
        pass
    try:

        if request.method == 'POST' and user_status_field.validate():
            return redirect(url_for('dashboarduser_categorysearch',keyword = user_status_field.category.data))
    except:
        pass
    return render_template('Operatordashboard_users.html',form = user_search_field,customers_list = customers_list,form2 = user_status_field)


@app.route('/dashboard/users/search=<keyword>',methods=['GET', 'POST'])
def dashboarduser_usernamesearch(keyword):
    user_search_field = SearchUserField(request.form)
    user_status_field = SearchUserStatus(request.form)
    dbmain = shelve.open('main.db','c')
    customers_dict = {}
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    customers_list = []
    for key in customers_dict:
        customer = customers_dict.get(key)
        if keyword in customer.get_username():
            customers_list.append(customer)
        else:
            pass

    #search by name
    try:
        if request.method == 'POST' and user_search_field.validate():
            return redirect(url_for('dashboarduser_usernamesearch',keyword = user_search_field.searchfield.data))
    except:
        pass

    #search by status
    try:

        if request.method == 'POST' and user_status_field.validate():
            return redirect(url_for('dashboarduser_categorysearch',keyword = user_status_field.category.data))
    except:
        pass
    

    
    
    
    return render_template('Operatordashboard_users_search.html',form = user_search_field,customers_list = customers_list,form2 = user_status_field,searchcondition = keyword)
    
@app.route('/dashboard/users/status=<keyword>',methods = ['GET','POST'])
def dashboarduser_categorysearch(keyword):
    user_search_field = SearchUserField(request.form)
    user_status_field = SearchUserStatus(request.form)
    dbmain = shelve.open('main.db','c')
    customers_dict = {}
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")
    pass

    customers_list = []
    for key in customers_dict:
        customer = customers_dict.get(key)
        if keyword == customer.get_status():
            customers_list.append(customer)
        else:
            pass

    #search by name
    try:
        if request.method == 'POST' and user_search_field.validate():
            return redirect(url_for('dashboarduser_usernamesearch',keyword = user_search_field.searchfield.data))
    except:
        pass

    #search by status
    try:

        if request.method == 'POST' and user_status_field.validate():
            return redirect(url_for('dashboarduser_categorysearch',keyword = user_status_field.category.data))
    except:
        pass
    
    return render_template('Operatordashboard_users_category.html',form = user_search_field,customers_list = customers_list,form2 = user_status_field,searchcondition = keyword)


@app.route('/operatorviewprofile/<int:id>',methods=['GET', 'POST'])
def operatorviewprofile(id): #id of profile they are viewing
    dbmain = shelve.open('main.db','c')
    
    customers_dict = {} #local one
    listings_dict = {}
    operatoractions_dict = {}
    terminate_user = OperatorTerminateUser(request.form,typeofaction='terminate user')
    suspend_user = OperatorSuspendUser(request.form,typeofaction='suspend user')
    restore_user = OperatorRestoreUser(request.form,typeofaction='restore user')
    restore_listing = OperatorRestoreListing(request.form,typeofaction='restore listing')
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    #sync listing dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Listing.Listing.count_ID = dbmain["ListingsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main count or count is at 0")

    #sync db5
    try:
        if "operatoractions" in dbmain:
            operatoractions_dict = dbmain["operatoractions"] #sync local with db1
        else:
            dbmain['operatoractions'] = operatoractions_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        operatoractions.Operatoractions.count_ID = dbmain["operatoractionsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main count or count is at 0")
    
    #get ID list of current user listings
    customer = customers_dict.get(id)
    customer_listings = customer.get_listings()
    print(f"\n*start of message *\nCurrent user has the following listings:{customer_listings}\n*end of message*")
    listing_list = []

    if customer.get_status() == "active": #show listings only when not suspended/terminated
        for key in listings_dict:
            print(key)
            if key in customer_listings:
                listing = listings_dict.get(key)
                listing_list.append(listing)
    
    #terminate user func
    if request.method == 'POST' and terminate_user.validate():
        if terminate_user.password.data == "sysadmin1":
            #create operator action object
            operator_action = operatoractions.Operatoractions(id,terminate_user.typeofaction.data,terminate_user.category.data,terminate_user.terminate_text.data)
            operatoractions_dict[operator_action.get_ID()] = operator_action #store into local

            #syncs db5 with local dict
            #syncs db5 count with local count (aka customer class)
            dbmain['operatoractions'] = operatoractions_dict
            dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID

            #make changes to affected user
            customer = customers_dict.get(id)
            customer.set_status("terminated")
            dbmain['Customers'] = customers_dict
            return(redirect(url_for('operatorviewprofile', id=id)))
        
    #restore user func
    if request.method == 'POST' and restore_user.validate():
        if restore_user.password.data == "sysadmin1":
            #create operator action object
            operator_action = operatoractions.Operatoractions(id,restore_user.typeofaction.data,"nil","nil")
            operatoractions_dict[operator_action.get_ID()] = operator_action #store into local

            #syncs db5 with local dict
            #syncs db5 count with local count (aka customer class)
            dbmain['operatoractions'] = operatoractions_dict
            dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID

            #make changes to affected user
            customer = customers_dict.get(id)
            customer.set_status("active")
            dbmain['Customers'] = customers_dict
            return(redirect(url_for('operatorviewprofile', id=id)))
    
    return render_template('OperatorViewProfile.html', customer=customer,
                            listings_list = listing_list,terminate_user_form = terminate_user, suspend_user_form = suspend_user, restore_user_form = restore_user,)

@app.route('/operatorviewprofilereviews/<int:id>',methods=['GET', 'POST'])
def operatorviewprofilereviews(id):
    dbmain = shelve.open('main.db','c')

    reviews_dict ={}
    customers_dict = {} #local one
    operatoractions_dict = {}
    terminate_user = OperatorTerminateUser(request.form,typeofaction='terminate user')
    suspend_user = OperatorSuspendUser(request.form,typeofaction='suspend user')
    restore_user = OperatorRestoreUser(request.form,typeofaction='restore user')
    #make sure local and db1 are the same state
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main customer count or count is at 0")

    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Reviews" in dbmain:
            reviews_dict = dbmain["Reviews"] #sync local with db1
        else:
            dbmain['Reviews'] = reviews_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Reviews.Reviews.count_ID = dbmain["ReviewsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Review count or count is at 0")

    #sync db5
    try:
        if "operatoractions" in dbmain:
            operatoractions_dict = dbmain["operatoractions"] #sync local with db1
        else:
            dbmain['operatoractions'] = operatoractions_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        operatoractions.Operatoractions.count_ID = dbmain["operatoractionsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB5 operatoractions count or count is at 0")

    customer = customers_dict.get(id)
    customer_reviews = customer.get_reviews()#return list of review IDs
    print(customer_reviews)
    customer_reviews_list = [] #THIS is the one sent to the html 
    
    for key in reviews_dict:
        print(key)
        if key in customer_reviews:
            print(key)
            review = reviews_dict.get(key)
            customer_reviews_list.append(review)


    
    #suspend user func
    if request.method == 'POST' and suspend_user.validate():
        if suspend_user.password.data == "sysadmin1":
            #create operator action object
            operator_action = operatoractions.Operatoractions(id,suspend_user.typeofaction.data,suspend_user.category.data,suspend_user.suspend_text.data)
            operatoractions_dict[operator_action.get_ID()] = operator_action #store into local

            #syncs db5 with local dict
            #syncs db5 count with local count (aka customer class)
            dbmain['operatoractions'] = operatoractions_dict
            dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID
            

            #make changes to affected user
            customer = customers_dict.get(id)
            customer.set_status("suspended")
            dbmain['Customers'] = customers_dict
            return(redirect(url_for('operatorviewprofile', id=id)))
        
    
    #terminate user func
    if request.method == 'POST' and terminate_user.validate():
        if terminate_user.password.data == "sysadmin1":
            #create operator action object
            operator_action = operatoractions.Operatoractions(id,terminate_user.typeofaction.data,terminate_user.category.data,terminate_user.terminate_text.data)
            operatoractions_dict[operator_action.get_ID()] = operator_action #store into local

            #syncs db5 with local dict
            #syncs db5 count with local count (aka customer class)
            dbmain['operatoractions'] = operatoractions_dict
            dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID


            #make changes to affected user
            customer = customers_dict.get(id)
            customer.set_status("terminated")
            dbmain['Customers'] = customers_dict
            return(redirect(url_for('operatorviewprofile', id=id)))
    
        
    #restore user func
    if request.method == 'POST' and restore_user.validate():
        if restore_user.password.data == "sysadmin1":
            #create operator action object
            operator_action = operatoractions.Operatoractions(id,restore_user.typeofaction.data,"nil","nil")
            operatoractions_dict[operator_action.get_ID()] = operator_action #store into local

            #syncs db5 with local dict
            #syncs db5 count with local count (aka customer class)
            dbmain['operatoractions'] = operatoractions_dict
            dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID

            #make changes to affected user
            customer = customers_dict.get(id)
            customer.set_status("active")
            dbmain['Customers'] = customers_dict
            return(redirect(url_for('operatorviewprofile', id=id)))
    

    return render_template("OperatorViewProfile_reviews.html",customer = customer ,number_of_reviews = len(customer_reviews_list),list_reviews = customer_reviews_list,terminate_user_form = terminate_user, suspend_user_form = suspend_user, restore_user_form = restore_user)

@app.route('/operatorviewprofilefeedback/<int:id>',methods=['GET', 'POST'])
def operatorviewprofilefeedback(id):
    dbmain = shelve.open('main.db', 'c')
    feedbacks_dict = {}
    customers_dict = {}  # local one
    operatoractions_dict = {}
    terminate_user = OperatorTerminateUser(request.form, typeofaction='terminate user')
    suspend_user = OperatorSuspendUser(request.form, typeofaction='suspend user')
    restore_user = OperatorRestoreUser(request.form, typeofaction='restore user')

    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"]  # sync local with db1
        else:
            dbmain['Customers'] = customers_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening customer.db")

    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        Customer.Customer.count_id = dbmain["CustomerCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    try:
        if "Feedback" in dbmain:
            feedbacks_dict = dbmain["Feedback"] #sync local with db1
        else:
            dbmain['Feedback'] = feedbacks_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        Feedback.Feedback.count_ID = dbmain["FeedbackCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Feedback count or count is at 0")
    # sync db5
    try:
        if "operatoractions" in dbmain:
            operatoractions_dict = dbmain["operatoractions"]  # sync local with db1
        else:
            dbmain['operatoractions'] = operatoractions_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        operatoractions.Operatoractions.count_ID = dbmain["operatoractionsCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main operatoractions count or count is at 0")

    #get current customer
    customer = customers_dict.get(id)
    customer_feedbacks_list = customer.get_feedbacks()
    feedbacks_list = []
    for key in feedbacks_dict:
        if key in customer_feedbacks_list: #ensure is own customer
            feedback = feedbacks_dict.get(key)
            feedbacks_list.append(feedback)
            if not hasattr(feedback, "_Feedback__category"):
                feedback._Feedback__category = "General"

            
    numberfeedbacks = len(feedbacks_list)
    # suspend user func
    if request.method == 'POST' and suspend_user.validate():
        if suspend_user.password.data == "sysadmin1":
            # create operator action object
            operator_action = operatoractions.Operatoractions(id, suspend_user.typeofaction.data,
                                                                  suspend_user.category.data,
                                                                  suspend_user.suspend_text.data)
            operatoractions_dict[operator_action.get_ID()] = operator_action  # store into local

            # syncs db5 with local dict
            # syncs db5 count with local count (aka customer class)
            dbmain['operatoractions'] = operatoractions_dict
            dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID
            

            # make changes to affected user
            customer = customers_dict.get(id)
            customer.set_status("suspended")
            dbmain['Customers'] = customers_dict
            return (redirect(url_for('operatorviewprofile', id=id)))

    if request.method == 'POST' and terminate_user.validate():
        if terminate_user.password.data == "sysadmin1":
            # create operator action object
            operator_action = operatoractions.Operatoractions(id, terminate_user.typeofaction.data,
                                                              terminate_user.category.data,
                                                              terminate_user.terminate_text.data)
            operatoractions_dict[operator_action.get_ID()] = operator_action  # store into local

            # syncs db5 with local dict
            # syncs db5 count with local count (aka customer class)
            dbmain['operatoractions'] = operatoractions_dict
            dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID

            # make changes to affected user
            customer = customers_dict.get(id)
            customer.set_status("terminated")
            dbmain['Customers'] = customers_dict
            return (redirect(url_for('operatorviewprofile', id=id)))

    # restore user func
    if request.method == 'POST' and restore_user.validate():
        if restore_user.password.data == "sysadmin1":
            # create operator action object
            operator_action = operatoractions.Operatoractions(id, restore_user.typeofaction.data, "nil", "nil")
            operatoractions_dict[operator_action.get_ID()] = operator_action  # store into local

            # syncs db5 with local dict
            # syncs db5 count with local count (aka customer class)
            dbmain['operatoractions'] = operatoractions_dict
            dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID
        

            # make changes to affected user
            customer = customers_dict.get(id)
            customer.set_status("active")
            dbmain['Customers'] = customers_dict
            return (redirect(url_for('operatorviewprofile', id=id)))

    return render_template("OperatorViewProfileFeedback.html",number_of_feedbacks =numberfeedbacks, customer=customer,feedbacks_list=feedbacks_list,terminate_user_form=terminate_user, suspend_user_form=suspend_user,restore_user_form=restore_user)


#operator actions

#listing related
@app.route('/invalidlisting')
def invalidlisting():
    return render_template('Operatorinvalidlisting.html')

#opstats - listing
@app.route('/operatordisablelisting/<int:listingid>',methods=['GET', 'POST'])
def operatordisablelisting(listingid):
    dbmain = shelve.open('main.db','c')
    operatoractions_dict = {}
    listings_dict = {}
    if listingid != 0:

        disable_listing = OperatorDisableListing(request.form,typeofaction='disable listing',listingid=listingid)
    else:
        disable_listing = OperatorDisableListing(request.form,typeofaction='disable listing')
    
    #sync listing dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Listing.Listing.count_ID = dbmain["ListingsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main count or count is at 0")

    #sync db5
    try:
        if "operatoractions" in dbmain:
            operatoractions_dict = dbmain["operatoractions"] #sync local with db1
        else:
            dbmain['operatoractions'] = operatoractions_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        operatoractions.Operatoractions.count_ID = dbmain["operatoractionsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main count or count is at 0")

    
    if request.method == 'POST' and disable_listing.validate():
        if disable_listing.password.data == "sysadmin1":
            listing = listings_dict.get(int(disable_listing.listingid.data))
            if (int(disable_listing.listingid.data)) in listings_dict and listing.get_status() == "available":


                #create operator action object)
                operator_action = operatoractions.Operatoractions(id,disable_listing.typeofaction.data,disable_listing.category.data,disable_listing.disable_text.data)
                operatoractions_dict[operator_action.get_ID()] = operator_action #store into local
                
                #syncs db5 with local dict
                #syncs db5 count with local count (aka customer class)
                dbmain['operatoractions'] = operatoractions_dict
                dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID


                #make changes to affected listing
                listing = listings_dict.get(int(disable_listing.listingid.data))
                listing.set_status("disabled")
                print(listing.get_status())
                dbmain['Listings'] = listings_dict

                try:
                    Operatorstats.operatorstats_listings("available","minus")
                    Operatorstats.operatorstats_listings("disabled","plus")
                except:
                    pass
                return(redirect(url_for('operatorcontrolcenter')))
            else:
                return (redirect(url_for('invalidlisting')))
        else:
            print("wrong password")

    return render_template('OperatorDisableListing.html',form = disable_listing)

#opstats - listing
@app.route('/operatorrestorelisting/<int:listingid>',methods=['GET','POST'])
def operatorrestorelisting(listingid):
    dbmain = shelve.open('main.db','c')
    operatoractions_dict = {}
    listings_dict = {}
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Listing.Listing.count_ID = dbmain["ListingsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main count or count is at 0")
    if listingid != 0:

        restore_listing = OperatorRestoreListing(request.form,typeofaction='restore user',listingid=listingid)
    else:
        restore_listing = OperatorRestoreListing(request.form,typeofaction='restore user')

    
    #restore listing
    if request.method == 'POST' and restore_listing.validate():
        if restore_listing.password.data == "sysadmin1":
            listing = listings_dict.get(int(restore_listing.listingid.data))
            if (int(restore_listing.listingid.data)) in listings_dict and listing.get_status() != "available":
                #create operator action object
                operator_action = operatoractions.Operatoractions(id,restore_listing.typeofaction.data,"nil","nil")
                operatoractions_dict[operator_action.get_ID()] = operator_action #store into local
                
                #syncs db5 with local dict
                #syncs db5 count with local count (aka customer class)
                dbmain['operatoractions'] = operatoractions_dict
                dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID

                #make changes to affected listing
                listing = listings_dict.get(int(restore_listing.listingid.data))
                listing.set_status("available")
                print(listing.get_status())
                dbmain['Listings'] = listings_dict
                try:
                    Operatorstats.operatorstats_listings("available","plus")
                    Operatorstats.operatorstats_listings("disabled","minus")
                except:
                    pass
                return(redirect(url_for('operatorcontrolcenter')))
            else:
                return (redirect(url_for('invalidlisting')))
        else:
            print("invalid password")
    return render_template('OperatorRestoreListing.html',form = restore_listing)

#invalid user page
@app.route('/invaliduser')
def invaliduser():
    return render_template('Operatorinvaliduser.html')

#user already suspended
@app.route('/useralreadysuspended')
def useralreadysuspended():
    return render_template('Useralreadysuspended.html')

#user already terminated
@app.route('/useralreadyterminated')
def useralreadyterminated():
    return render_template('Useralreadyterminated.html')

#suspend user
@app.route('/operatorsuspenduser/<int:customerid>',methods=['GET','POST']) #page to enter details
def operatorsuspenduser(customerid):
    customers_dict = {}
    suspend_user = OperatorSuspendUser(request.form,typeofaction='suspend user',affectedid=customerid)
    dbmain = shelve.open('main.db','c')
    #prevent data being outdated
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    if request.method == 'POST' and suspend_user.validate():
        if suspend_user.password.data == "sysadmin1":
            #check if ID exists
            flag = False
            for key in customers_dict:
                if key == int(suspend_user.affectedid.data):
                    flag = True
                    break
                else:
                    flag = False
            customer = customers_dict.get(int(suspend_user.affectedid.data))
            if flag == True:
                if customer.get_status() == "active":

                    session['typeofaction'] = suspend_user.typeofaction.data
                    session['category'] = suspend_user.category.data
                    session['suspend_text'] = suspend_user.suspend_text.data
                    return(redirect(url_for('confirmsuspenduser',customerid=int(suspend_user.affectedid.data))))
                elif customer.get_status() == "suspended":
                    return(redirect(url_for('useralreadysuspended')))
                elif customer.get_status() == "terminated":
                    return(redirect(url_for('useralreadyterminated')))
            else:
                return(redirect(url_for('invaliduser')))
    else:
        pass
            
    return render_template('Operatorsuspenduser.html',form=suspend_user,customerID = customerid)

#measure in place prevent accidentally suspending
@app.route('/confirmsuspenduser/<int:customerid>',methods=['GET','POST'])#displays 2 options, confirm or go back
def confirmsuspenduser(customerid):
    customers_dict = {}
    dbmain = shelve.open('main.db','c')
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    
    #retrive customer obj
    customer = customers_dict.get(customerid)

    return render_template('Operatorconfirmsuspenduser.html',customer=customer)

#actual code, opstats - user
@app.route('/suspenduser/<int:customerid>')
def suspenduser(customerid):
    dbmain = shelve.open('main.db','c')
    operatoractions_dict = {}
    customers_dict = {}
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")
    try:
        if "operatoractions" in dbmain:
            operatoractions_dict = dbmain["operatoractions"] #sync local with db1
        else:
            dbmain['operatoractions'] = operatoractions_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        operatoractions.Operatoractions.count_ID = dbmain["operatoractionsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main count or count is at 0")
    #create operator action object
    operator_action = operatoractions.Operatoractions(customerid,session['typeofaction'],session['category'],session['suspend_text'])
    operatoractions_dict[operator_action.get_ID()] = operator_action #store into local

    #syncs db5 with local dict
    #syncs db5 count with local count (aka customer class)
    dbmain['operatoractions'] = operatoractions_dict
    dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID

    try:
        Operatorstats.operatorstats_opactions("total","plus")
    except:
        pass
    #make changes to affected user
    customer = customers_dict.get(customerid)
    customer.set_status("suspended")
    dbmain['Customers'] = customers_dict

    try:
        Operatorstats.operatorstats_users("suspended","plus")
        Operatorstats.operatorstats_users("active","minus")
    except:
        print("Error! Operator stats did not update")

    return(redirect(url_for('operatorviewprofile', id=customerid)))


#terminate user
@app.route('/operatorterminateuser/<int:customerid>',methods=['GET','POST'])
def operatorterminateuser(customerid):
    customers_dict = {}
    terminate_user = OperatorTerminateUser(request.form,typeofaction='terminate user',affectedid=customerid)
    dbmain = shelve.open('main.db','c')
    #prevent data being outdated
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    if request.method == 'POST' and terminate_user.validate():
        if terminate_user.password.data == "sysadmin1":
            #check if ID exists
            flag = False
            for key in customers_dict:
                if key == int(terminate_user.affectedid.data):
                    flag = True
                    break
                else:
                    flag = False
            customer = customers_dict.get(int(terminate_user.affectedid.data))
            if flag == True:
                if customer.get_status() == "active":
                    session['typeofaction'] = terminate_user.typeofaction.data
                    session['category'] = terminate_user.category.data
                    session['terminate_text'] = terminate_user.terminate_text.data
                    return(redirect(url_for('confirmterminateuser',customerid=int(terminate_user.affectedid.data))))
                elif customer.get_status() == "suspended":
                    return(redirect(url_for('useralreadysuspended')))
                elif customer.get_status() == "terminated":
                    return(redirect(url_for('useralreadyterminated')))
            else:
                return(redirect(url_for('invaliduser')))
    else:
        pass
            
    return render_template('Operatorterminateuser.html',form=terminate_user,customerID = customerid)

@app.route('/confirmterminateuser/<int:customerid>',methods=['GET','POST'])
def confirmterminateuser(customerid):
    customers_dict = {}
    dbmain = shelve.open('main.db','c')
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    
    #retrive customer obj
    customer = customers_dict.get(customerid)
    return render_template('Operatorconfirmterminateuser.html',customer=customer)

#opstats - terminate user
@app.route('/terminateuser/<int:customerid>',methods=['GET','POST'])
def terminateuser(customerid):
    dbmain = shelve.open('main.db','c')
    operatoractions_dict = {}
    customers_dict = {}
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")
    try:
        if "operatoractions" in dbmain:
            operatoractions_dict = dbmain["operatoractions"] #sync local with db1
        else:
            dbmain['operatoractions'] = operatoractions_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        operatoractions.Operatoractions.count_ID = dbmain["operatoractionsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main count or count is at 0")
    #create operator action object
    operator_action = operatoractions.Operatoractions(customerid,session['typeofaction'],session['category'],session['terminate_text'])
    operatoractions_dict[operator_action.get_ID()] = operator_action #store into local

    #syncs db5 with local dict
    #syncs db5 count with local count (aka customer class)
    dbmain['operatoractions'] = operatoractions_dict
    dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID
    try:
        Operatorstats.operatorstats_opactions("total","plus")
    except:
        pass
    #make changes to affected user
    customer = customers_dict.get(customerid)
    customer.set_status("terminated")
    dbmain['Customers'] = customers_dict

    try:
        Operatorstats.operatorstats_users("terminated","plus")
        Operatorstats.operatorstats_users("active","minus")
    except:
        print("Error! Operator stats did not update")

    return(redirect(url_for('operatorviewprofile', id=customerid)))

#restore user
@app.route('/operatorrestoreuser/<int:customerid>',methods=['GET','POST'])
def operatorrestoreuser(customerid):
    customers_dict = {}
    restore_user = OperatorRestoreUser(request.form,typeofaction='restore user',affectedid=customerid)
    dbmain = shelve.open('main.db','c')
    #prevent data being outdated
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    if request.method == 'POST' and restore_user.validate():
        if restore_user.password.data == "sysadmin1":
            #check if ID exists
            flag = False
            for key in customers_dict:
                if key == int(restore_user.affectedid.data):
                    flag = True
                    break
                else:
                    flag = False
            
            if flag == True:
                session['typeofaction'] = restore_user.typeofaction.data
                session['category'] = "nil"
                session['restore_text'] = "nil"
                return(redirect(url_for('confirmrestoreuser',customerid=int(restore_user.affectedid.data))))
            else:
                return(redirect(url_for('invaliduser')))
    else:
        print("restore user redirect no work")
            
    return render_template('Operatorrestoreuser.html',form=restore_user,customerID = customerid)

@app.route('/confirmrestoreuser/<int:customerid>',methods=['GET','POST'])
def confirmrestoreuser(customerid):
    customers_dict = {}
    dbmain = shelve.open('main.db','c')
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    
    #retrive customer obj
    customer = customers_dict.get(customerid)
    return render_template('Operatorconfirmrestoreuser.html',customer=customer)

#opstats - user
@app.route('/restoreuser/<int:customerid>',methods=['GET','POST'])
def restoreuser(customerid):
    dbmain = shelve.open('main.db','c')
    operatoractions_dict = {}
    customers_dict = {}
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Customer.Customer.count_id = dbmain["CustomerCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")
    try:
        if "operatoractions" in dbmain:
            operatoractions_dict = dbmain["operatoractions"] #sync local with db1
        else:
            dbmain['operatoractions'] = operatoractions_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        operatoractions.Operatoractions.count_ID = dbmain["operatoractionsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main count or count is at 0")
    #create operator action object
    operator_action = operatoractions.Operatoractions(customerid,session['typeofaction'],session['category'],session['restore_text'])
    operatoractions_dict[operator_action.get_ID()] = operator_action #store into local
    try:
        Operatorstats.operatorstats_opactions("total","plus")
    except:
        pass
    #syncs db5 with local dict
    #syncs db5 count with local count (aka customer class)
    dbmain['operatoractions'] = operatoractions_dict
    dbmain['operatoractionsCount'] = operatoractions.Operatoractions.count_ID

    #make changes to affected user
    customer = customers_dict.get(customerid)

    if customer.get_status() == "suspended":
        try:
            Operatorstats.operatorstats_users("suspended","minus")
            Operatorstats.operatorstats_users("active","plus")
        except:
            print("Error! Operator stats did not update")
    elif customer.get_status() == "terminated":
        try:
            Operatorstats.operatorstats_users("terminated","minus")
            Operatorstats.operatorstats_users("active","plus")
        except:
            print("Error! Operator stats did not update")
    else:
        print("Error! Can't find customer status under restore user")

    customer.set_status("active")
    dbmain['Customers'] = customers_dict
        
    return(redirect(url_for('operatorviewprofile', id=customerid)))

@app.route('/dashboard/listings', methods=['GET', 'POST'])
def dashboardlistings():
    listing_search_field = SearchListingField(request.form)
    listing_searchstatus_field = SearchListingStatusField(request.form)
    listing_searchid_field = SearchListingIDField(request.form)
    dbmain = shelve.open('main.db', 'c')
    listings_dict = {}

    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"]  # Sync local with db
        else:
            dbmain['Listings'] = listings_dict  # Initialize db if empty
    except Exception as e:
        print("Error in opening main.db:", e)

    listings_list = list(listings_dict.values())  # Convert dict to list

    # Detect which form was submitted
    form_action = request.form.get("action")

    if request.method == 'POST':
        if form_action == "generate_report":
            return generate_excel_report(listings_list)
    try:
        if request.method == 'POST' and listing_search_field.validate():
            return redirect(url_for('dashboardlistingssearch', keyword=listing_search_field.searchfield.data))
    except:
        pass

    try:
        if request.method == 'POST' and listing_searchid_field.validate():
            return redirect(url_for('dashboardlistingssearchid', id=listing_searchid_field.searchidfield.data))
    except:
        pass

    try:
        if request.method == 'POST' and listing_searchstatus_field.validate():
            return redirect(
                url_for('dashboardlistingssearchstatus', status=listing_searchstatus_field.searchstatusfield.data))
    except:
        pass
    return render_template(
        'Operatordashboard_listings.html',
        form=listing_search_field,
        form2=listing_searchid_field,
        form3=listing_searchstatus_field,
        listings_list=listings_list
    )
def generate_excel_report(listings_list):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listings Report"

    # Define headers (adjust based on your listing attributes)
    headers = ["Listing ID", "Creator username", "Creator ID", "Title", "Description", "Category", "Status", "Creation Date", "Likes", "Sold date"]
    ws.append(headers)

    # Bold the header row
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Append data rows from listings_list
    for listing in listings_list:
        # If listing is an object, use its attributes; if it's a dict, adjust accordingly.
        row = [
            listing.get_ID(),
            listing.get_creator_username(),
            listing.get_creatorID(),
            listing.get_title(),
            listing.get_description(),
            listing.get_category(),
            listing.get_status(),
            listing.get_creation_date(),
            listing.get_likes(),
            listing.get_soldDate()
        ]
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(wrap_text=True)
    for col_num in range(1, ws.max_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name="listings_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/dashboard/listings/search=<keyword>',methods=['GET', 'POST'])
def dashboardlistingssearch(keyword):
    listing_search_field = SearchListingField(request.form)
    listing_searchstatus_field = SearchListingStatusField(request.form)
    listing_searchid_field = SearchListingIDField(request.form)
    dbmain = shelve.open('main.db','c')
    listings_dict = {}

    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    listings_list = []        
    for key in listings_dict:
        listing = listings_dict.get(key)
        if keyword in listing.get_title():
            listings_list.append(listing)
        else:
            pass

    try:
        if request.method == 'POST' and listing_search_field.validate():
            return redirect(url_for('dashboardlistingssearch',keyword = listing_search_field.searchfield.data))
    except:
        pass

    try:
        if request.method == 'POST' and listing_searchid_field.validate():
            return redirect(url_for('dashboardlistingssearchid',id = listing_searchid_field.searchidfield.data))
    except:
        pass

    try:
        if request.method == 'POST' and listing_searchstatus_field.validate():
            return redirect(url_for('dashboardlistingssearchstatus',status = listing_searchstatus_field.searchstatusfield.data))
    except:
        pass
    
    return render_template('Operatordashboard_listings_search.html',form = listing_search_field, form2 = listing_searchid_field,form3 = listing_searchstatus_field,listings_list = listings_list,searchcondition=keyword)

@app.route('/dashboard/listings/id=<id>',methods = ['GET','POST'])
def dashboardlistingssearchid(id):
    listing_search_field = SearchListingField(request.form)
    listing_searchstatus_field = SearchListingStatusField(request.form)
    listing_searchid_field = SearchListingIDField(request.form)
    dbmain = shelve.open('main.db','c')
    listings_dict = {}
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    listings_list = []        
    for key in listings_dict:
        if key == id:
            listing = listings_list.get(key)
            listings_list.append(listing)
            break
    try:
        if request.method == 'POST' and listing_search_field.validate():
            return redirect(url_for('dashboardlistingssearch',keyword = listing_search_field.searchfield.data))
    except:
        pass

    try:
        if request.method == 'POST' and listing_searchid_field.validate():
            return redirect(url_for('dashboardlistingssearchid',id = listing_searchid_field.searchidfield.data))
    except:
        pass

    try:
        if request.method == 'POST' and listing_searchstatus_field.validate():
            return redirect(url_for('dashboardlistingssearchstatus',status = listing_searchstatus_field.searchstatusfield.data))
    except:
        pass
    
    return render_template('Operatordashboard_listings_searchid.html',form = listing_search_field, form2 = listing_searchid_field,form3 = listing_searchstatus_field,listings_list = listings_list,searchcondition = id)

@app.route('/dashboard/listings/status=<status>',methods = ['GET','POST'])
def dashboardlistingssearchstatus(status):
    listing_search_field = SearchListingField(request.form)
    listing_searchstatus_field = SearchListingStatusField(request.form)
    listing_searchid_field = SearchListingIDField(request.form)
    dbmain = shelve.open('main.db','c')
    listings_dict = {}

    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    listings_list = []        
    for key in listings_dict:
        listing = listings_dict.get(key)
        if listing.get_status() == status:
            listings_list.append(listing)
    try:
        if request.method == 'POST' and listing_search_field.validate():
            return redirect(url_for('dashboardlistingssearch',keyword = listing_search_field.searchfield.data))
    except:
        pass

    try:
        if request.method == 'POST' and listing_searchid_field.validate():
            return redirect(url_for('dashboardlistingssearchid',id = listing_searchid_field.searchidfield.data))
    except:
        pass

    try:
        if request.method == 'POST' and listing_searchstatus_field.validate():
            return redirect(url_for('dashboardlistingssearchstatus',status = listing_searchstatus_field.searchstatusfield.data))
    except:
        pass
    
    return render_template('Operatordashboard_listings_searchstatus.html',form = listing_search_field, form2 = listing_searchid_field,form3 = listing_searchstatus_field,listings_list = listings_list,searchcondition=status)

@app.route('/operatorviewlisting/<int:id>',)
def operatorviewlisting(id):
    dbmain = shelve.open('main.db','c')
    listings_dict = {}
    customers_dict = {}
    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"] #sync local with db2
        else:
            dbmain['Listings'] = listings_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"] #sync local with db1
        else:
            dbmain['Customers'] = customers_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    
    listing = listings_dict.get(id)
    for key in customers_dict:
        if key == listing.get_creatorID():
            seller = customers_dict.get(key)
            break
    
    return render_template('OperatorViewListing.html', listing = listing,seller = seller, current_sessionID = session_ID)
    
@app.route('/dashboard/reports',methods=['GET', 'POST'])
def dashboardreports():
    dbmain =shelve.open('main.db','c')
    reports_dict = {}
    report_search_field = SearchReportField(request.form)
    #sync report dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Reports" in dbmain:
            reports_dict = dbmain["Reports"] #sync local with db2
        else:
            dbmain['Reports'] = reports_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Report.Report.count_ID = dbmain["ReportsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Report count or count is at 0")

    reports_list = []
    for key in reports_dict:
        report = reports_dict.get(key)
        reports_list.append(report)
    try:
        if request.method == 'POST' and report_search_field.validate():
            return redirect(url_for('dashboardreportssearch',keyword = report_search_field.searchfield.data))
    except:
        pass
    return render_template("Operatordashboard_reports.html",reports_list=reports_list,form =report_search_field)

@app.route('/dashboard/reports/search=<keyword>',methods=['GET', 'POST'])
def dashboardreportssearch(keyword):
    dbmain =shelve.open('main.db','c')
    reports_dict = {}
    report_search_field = SearchReportField(request.form)
    #sync report dbs
    #PS JUST COPY AND PASTE IF YOU'RE ACCESSING IT
    try:
        if "Reports" in dbmain:
            reports_dict = dbmain["Reports"] #sync local with db2
        else:
            dbmain['Reports'] = reports_dict #sync db2 with local (basically null)
    except:
            print("Error in opening main.db")
    #sync listing IDs
    try:
        dbmain = shelve.open('main.db','c')    
        Report.Report.count_ID = dbmain["ReportsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main Report count or count is at 0")

    reports_list = []
    for key in reports_dict:
        report = reports_dict.get(key)
        if keyword in report.get_offender_username():
            reports_list.append(report)
    
    try:
        if request.method == 'POST' and report_search_field.validate():
            return redirect(url_for('dashboardreportssearch',keyword = report_search_field.searchfield.data))
    except:
        pass
    return render_template('Operatordashboard_reports_search.html',reports_list=reports_list,form=report_search_field)

@app.route('/dashboard/operatoractions',methods=['GET', 'POST'])
def dashboardoperatoractions():
    dbmain = shelve.open('main.db','c')
    operatoractions_dict = {}
    operatoraction_search_field = SearchOperatorActionField(request.form)
    #sync db5
    try:
        if "operatoractions" in dbmain:
            operatoractions_dict = dbmain["operatoractions"] #sync local with db1
        else:
            dbmain['operatoractions'] = operatoractions_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        operatoractions.Operatoractions.count_ID = dbmain["operatoractionsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB main operatoractions count or count is at 0")
    operator_actions_list = []
    for key in operatoractions_dict:
        operatoraction = operatoractions_dict.get(key)
        operator_actions_list.append(operatoraction)
    try:
        if request.method == 'POST' and operatoraction_search_field.validate():
            return redirect(url_for('dashboardoperatoractionssearch',keyword = operatoraction_search_field.searchfield.data))
    except:
        pass
    return render_template('Operatordashboard_operatoraction.html',form = operatoraction_search_field,operator_actions_list=operator_actions_list)

@app.route('/dashboard/operatoractions/search=<keyword>',methods=['GET', 'POST'])
def dashboardoperatoractionssearch(keyword):
    dbmain = shelve.open('main.db','c')
    operatoractions_dict = {}
    operatoraction_search_field = SearchOperatorActionField(request.form)
    #sync db5
    try:
        if "operatoractions" in dbmain:
            operatoractions_dict = dbmain["operatoractions"] #sync local with db1
        else:
            dbmain['operatoractions'] = operatoractions_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
        
    #sync IDs
    try:
        dbmain = shelve.open('main.db','c')    
        operatoractions.Operatoractions.count_ID = dbmain["operatoractionsCount"] #sync count between local and db1
    except:
        print("Error in retrieving data from DB5 operatoractions count or count is at 0")

    operator_actions_list = []
    for key in operatoractions_dict:
        operatoraction = operatoractions_dict.get(key)
        if operatoraction.get_category() == keyword:
            operator_actions_list.append(operatoraction)
    try:
        if request.method == 'POST' and operatoraction_search_field.validate():
            return redirect(url_for('dashboardoperatoractionssearch',keyword = operatoraction_search_field.searchfield.data))
    except:
        pass
    
    return render_template('Operatordashboard_operatoraction_search.html',form = operatoraction_search_field,operator_actions_list=operator_actions_list)

@app.route('/dashboard/feedbacks', methods=['GET','POST'])
def dashboardfeedbacks():
    search_replied = FilterFeedback(request.form)
    feedbacks_dict = {}
    dbmain = shelve.open('main.db','c')
    try:
        if "Feedback" in dbmain:
            feedbacks_dict = dbmain["Feedback"] #sync local with db1
        else:
            dbmain['Feedback'] = feedbacks_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    feedbacks_list=[]
    for key in feedbacks_dict:
        feedback = feedbacks_dict.get(key)
        feedbacks_list.append(feedback)
        if not hasattr(feedback, "_Feedback__category"):
            feedback._Feedback__category = "General"
    if request.method == "POST" and search_replied.validate():
        return(redirect(url_for('dashboardfeedbackssearch',keyword = search_replied.searchstatusfield.data)))

    return render_template("Operatordashboard_feedback.html",feedbacks_list=feedbacks_list,form = search_replied)

@app.route('/dashboard/feedbacks/<keyword>',methods=['GET','POST'])
def dashboardfeedbackssearch(keyword):
    search_replied = FilterFeedback(request.form)
    feedbacks_dict = {}
    dbmain = shelve.open('main.db','c')
    try:
        if "Feedback" in dbmain:
            feedbacks_dict = dbmain["Feedback"] #sync local with db1
        else:
            dbmain['Feedback'] = feedbacks_dict #sync db1 with local (basically null)
    except:
        print("Error in opening main.db")
    feedbacks_list=[]
    for key in feedbacks_dict:
        feedback = feedbacks_dict.get(key)
        if keyword == "unreplied":
            if feedback.get_reply() == None:
                feedbacks_list.append(feedback)
            else:
                pass
        elif keyword == "replied":
            if feedback.get_reply() != None:
                feedbacks_list.append(feedback)
            else:
                pass
        else:
            pass
    
    if request.method == "POST" and search_replied.validate():
        return(redirect(url_for('dashboardfeedbackssearch',keyword = search_replied.searchstatusfield.data)))


    return render_template("Operatordashboard_feedback_search.html",feedbacks_list=feedbacks_list,form = search_replied,searchcondition = keyword)


@app.route('/dashboard/OperatorDashboard', methods=['GET', 'POST'])
def dashboard_dashboard():
    dbmain = shelve.open('main.db', 'c')
    listings_dict = {}
    customers_dict = {}
    try:
        if "Customers" in dbmain:
            customers_dict = dbmain["Customers"]  # sync local with db1
        else:
            dbmain['Customers'] = customers_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    # sync IDs
    try:
        dbmain = shelve.open('main.db', 'c')
        Customer.Customer.count_id = dbmain["CustomerCount"]  # sync count between local and db1
    except:
        print("Error in retrieving data from DB main Customer count or count is at 0")

    try:
        if "Listings" in dbmain:
            listings_dict = dbmain["Listings"]
        else:
            dbmain["Listings"] = listings_dict
    except:
        print("Error in opening Listings data.")

    cust_ratings_list = []
    feedbacks_dict = {}
    feedbacks_count = 0
    try:
        if "Feedback" in dbmain:
            feedbacks_dict = dbmain["Feedback"]  # sync local with db1
        else:
            dbmain['Feedback'] = feedbacks_dict  # sync db1 with local (basically null)
    except:
        print("Error in opening main.db")

    for key, value in feedbacks_dict.items():
        cust_ratings_list.append(int(feedbacks_dict[key].get_rating()))
        feedbacks_count += 1
    if feedbacks_count != 0:
        customer_satisfaction_gauge = f"{((sum(cust_ratings_list)/feedbacks_count)/5) * 100:.1f}"
    else:
        customer_satisfaction_gauge = 0
    customer_total_feedbacks = feedbacks_count
    customers_list = []
    for key in customers_dict:
        customer = customers_dict.get(key)
        customers_list.append(customer)
    cat_electronics = 0
    cat_books = 0
    cat_fashion = 0
    cat_entertainment = 0
    cat_misc = 0
    listings_list = list(listings_dict.values())
    seller_sales_count = {}
    listing_days = []
    top10_seller_dic = {}
    name_dic = {}
    rating_dic = {}
    cat_electronics_avg_days = []
    cat_books_avg_days = []
    cat_fashion_avg_days = []
    cat_entertainment_avg_days = []
    cat_misc_avg_days = []

    for listing in listings_list:

        category = listing.get_category()
        creatorID = listing.get_creatorID()
        username = listing.get_creator_username()
        customer11 = customers_dict.get(creatorID)
        rating = 0
        
        if customer11:
            rating = customer11.get_rating()
        # most sold categories
        if listing.get_status() == "sold":
            if category == "Category 1":
                cat_electronics += 1
            elif category == "Category 2":
                cat_books += 1
            elif category == "Category 3":
                cat_fashion += 1
            elif category == "Category 4":
                cat_entertainment += 1
            elif category == "Category 5":
                cat_misc += 1
        # top sellers
        if listing.get_status() == "sold":
            if creatorID in seller_sales_count:
                seller_sales_count[creatorID] += 1
            else:
                seller_sales_count[creatorID] = 1
        sorted_sellers = sorted(seller_sales_count.items(), key=lambda x: x[1], reverse=True)
        top_10_sellers_with_counts = sorted_sellers[:10]
        for seller_id, sold_count in top_10_sellers_with_counts:
            top10_seller_dic[seller_id] = sold_count
            if seller_id not in name_dic:
                name_dic[seller_id] = username
                rating_dic[seller_id] = rating

        # average number of days for listing to be sold
        if listing.get_status() == "sold":
            creationdate = listing.get_creation_date()
            solddate = listing.get_soldDate()
            date_format = "%d/%m/%y"
            creation_date = datetime.strptime(creationdate, date_format)
            sold_date = datetime.strptime(solddate, date_format)
            days_to_sell = (sold_date - creation_date).days
            listing_days.append(days_to_sell)
            if category == "Category 1":
                cat_electronics_avg_days.append(days_to_sell)
            elif category == "Category 2":
                cat_books_avg_days.append(days_to_sell)
            elif category == "Category 3":
                cat_fashion_avg_days.append(days_to_sell)
            elif category == "Category 4":
                cat_entertainment_avg_days.append(days_to_sell)
            elif category == "Category 5":
                cat_misc_avg_days.append(days_to_sell)

    sorted_top10_seller = sorted(top10_seller_dic.items(), key=lambda x: x[1], reverse=True)
    sorted_top10_seller = sorted_top10_seller[:10]
    sorted_top10_seller_dic = dict(sorted_top10_seller)
    if sum(listing_days) != 0:
        day_avg = sum(listing_days) / len(listing_days)
    else:
        day_avg = 0
    if sum(cat_electronics_avg_days) != 0:
        electronics_day_avg = sum(cat_electronics_avg_days) / len(
            cat_electronics_avg_days)  # amount of time for products to be sold
    else:
        electronics_day_avg = 0
    if sum(cat_books_avg_days) != 0:
        books_day_avg = sum(cat_books_avg_days) / len(cat_books_avg_days)
    else:
        books_day_avg = 0
    if sum(cat_fashion_avg_days) != 0:
        fashion_day_avg = sum(cat_fashion_avg_days) / len(cat_fashion_avg_days)
    else:
        fashion_day_avg = 0
    if sum(cat_entertainment_avg_days) != 0:
        entertainment_day_avg = sum(cat_entertainment_avg_days) / len(cat_entertainment_avg_days)
    else:
        entertainment_day_avg = 0
    if sum(cat_misc_avg_days) != 0:
        misc_day_avg = sum(cat_misc_avg_days) / len(cat_misc_avg_days)
    else:
        misc_day_avg = 0

    return render_template('OperatorDashboard.html', current_sessionID=session_ID, customer=customer,
                           avg_days_to_sell=day_avg, top10_seller_dic=sorted_top10_seller_dic,
                           cat_electronics=cat_electronics, cat_books=cat_books, cat_fashion=cat_fashion,
                           cat_entertainment=cat_entertainment, cat_misc=cat_misc, name_dic=name_dic,
                           rating_dic=rating_dic, electronics_day_avg=electronics_day_avg, books_day_avg=books_day_avg,
                           fashion_day_avg=fashion_day_avg, entertainment_day_avg=entertainment_day_avg,
                           misc_day_avg=misc_day_avg, customers_list=customers_list,
                           customer_satisfaction_gauge=customer_satisfaction_gauge,
                           customer_total_feedbacks = customer_total_feedbacks)


@app.route('/operator-dashboard', methods=['POST'])
def operator_dashboard():
    if request.method == 'POST':
        selected_options = request.form.getlist('options[]')
        user_id = request.form.get('user_id')
        print(f"ID: {user_id}, {selected_options}")
        # add selected options to an Excel file
        file_path = save_to_excel(selected_options, user_id)
        print("Reached")
        return send_file(file_path, as_attachment=True, download_name=f'User {user_id} account details.xlsx')
    return render_template('Operatordashboard_users.html')

@app.route('/operator-dashboard-download-all', methods=['POST'])
def operator_dashboard_download_all():
    if request.method == 'POST':
        dbmain = shelve.open('main.db', 'r')  # Open the database in read-only mode
        customers_dict = dbmain.get("Customers", {})
        users_list = list(customers_dict.values())
        return save_to_excel_all_user_reports(users_list)
def save_to_excel_all_user_reports(users_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Report"

        # Define headers (adjust based on your listing attributes)
    headers = ["User ID", "User username", "Email", "Rating", "Date joined", "Status"]
    ws.append(headers)

        # Bold the header row
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

        # Append data rows from listings_list
    for user in users_list:
            # If listing is an object, use its attributes; if it's a dict, adjust accordingly.
        row = [
            user.get_id(),
            user.get_username(),
            user.get_email(),
            user.get_rating(),
            user.get_date_joined(),
            user.get_status()
        ]
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(wrap_text=True)
    for col_num in range(1, ws.max_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name="users_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
def save_to_excel(selected_options, user_id):
    try:
        dbmain = shelve.open('main.db', 'r')  # Open the database in read-only mode
        customers_dict = dbmain.get("Customers", {})
        if int(user_id) not in customers_dict:
            print("User not found.")
        customer = customers_dict[int(user_id)]  # Fetch the customer by user_id
        # Retrieve the customer's data
        email_data = customer.get_email()
        username_data = customer.get_username()
        reports_dict = dbmain.get("Reports", {})
        reports_data = [report for report in reports_dict.values() if report.get_offender_ID() == int(user_id)]
        listings_dict = dbmain.get("Listings", {})
        listings_data = [listing for listing in listings_dict.values() if listing.get_creatorID() == int(user_id)]
        feedbacks_dict = {}
        try:
            if "Feedback" in dbmain:
                feedbacks_dict = dbmain["Feedback"]  # sync local with db1
            else:
                dbmain['Feedback'] = feedbacks_dict  # sync db1 with local (basically null)
        except:
            print("Error in opening main.db")
        review_dict = dbmain.get("Reviews", {})
        customer_reviews = customer.get_reviews()  # list of review IDs
        customer_reviews_list = []
        for key in review_dict:
            if key in customer_reviews:  # check if review ID matches customer's reviews
                review = review_dict.get(key)
                customer_reviews_list.append(review)  # append to list
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"User ID {user_id} Account Details"
        email_list = []
        username_list = []
        # Start with User ID header and data
        headers = [f"User ID: {user_id}"]

        if 'Email' in selected_options:
            headers.append("Email")
            email_list.append(email_data)  # Replace with actual email retrieval logic

        if 'Username' in selected_options:
            headers.append("Username")
            username_list.append(username_data)  # Replace with actual username retrieval logic

        if 'Report details' in selected_options:
            report_ws = wb.create_sheet("Report Details")  # Create a new sheet for reports
            report_ws.append(["Reporter ID", "Report category", "Report comments"])
            if reports_data:
                for report in reports_data:
                    report_ws.append([
                        report.get_creator_ID(),
                        report.get_category(),
                        report.get_comment()
                    ])
            else:
                report_ws.append([
                    "No reports available",
                    "No reports available",
                    "No reports available"
                ])
        if 'Listings information' in selected_options:
            listings_ws = wb.create_sheet("Listings information")
            listings_ws.append(["Listing ID", "Creator ID", "Name", "Description", "Condition", "Category", "Status"])
            if listings_data:
                for listing in listings_data:
                    listings_ws.append([
                        listing.get_ID(),
                        listing.get_creatorID(),
                        listing.get_title(),
                        listing.get_description(),
                        listing.get_condition(),
                        listing.get_category(),
                        listing.get_status()
                    ])
            else:
                listings_ws.append([
                    "No listings available",
                    "No listings available",
                    "No listings available",
                    "No listings available",
                    "No listings available",
                    "No listings available",
                    "No listings available"
                ])

        if 'Review information' in selected_options:
            review_ws = wb.create_sheet("Review information")
            review_ws.append(["Reviewer ID", "Rating", "Comment"])
            if customer_reviews_list:
                for review in customer_reviews_list:
                    review_ws.append([
                        review.get_creator_ID(),
                        review.get_rating(),
                        review.get_comment()
                    ])
            else:
                review_ws.append([
                    "No reviews available",
                    "No reviews available",
                    "No reviews available"
                ])

        if 'Feedbacks' in selected_options:
            review_ws = wb.create_sheet("Review information")
            review_ws.append(["Reviewer ID", "Rating", "Comment"])
            if customer_reviews_list:
                for review in customer_reviews_list:
                    review_ws.append([
                        review.get_creator_ID(),
                        review.get_rating(),
                        review.get_comment()
                    ])
            else:
                review_ws.append([
                    "No reviews available",
                    "No reviews available",
                    "No reviews available"
                ])
        # Append the headers row to the worksheet (row 1)
        ws.append(headers)

        # Determine the maximum number of rows needed (to ensure all columns have the same number of rows)
        max_rows = max(len(reports_data), 1)

        # Now, we need to append the data rows for each column
        for i in range(max_rows):
            row_data = ['']  # Always add the user ID as the first column

            if 'Email' in selected_options:
                # Add email data or empty string if no data for this row
                row_data.append(email_list[i] if i < len(email_list) else '')

            if 'Username' in selected_options:
                # Add username data or empty string if no data for this row
                row_data.append(username_list[i] if i < len(username_list) else '')
            # Append this row of data under the appropriate headers (starting from row 2)
            ws.append(row_data)

        # Adjust column widths based on the maximum content length (for headers and data)
        for ws in wb.sheetnames:
            sheet = wb[ws]

            # Iterate over all cells in the worksheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        # Apply text wrapping
                        cell.alignment = Alignment(wrap_text=True)

            # Set the fixed column width for all columns in the sheet
            for col_num in range(1, sheet.max_column + 1):
                column_letter = openpyxl.utils.get_column_letter(col_num)  # Convert column number to letter
                sheet.column_dimensions[column_letter].width = 20  # Set the column width to a fixed size
        # Save the file
        file_path = "selected_options.xlsx"
        wb.save(file_path)
        return file_path
    except Exception as e:
        print(f"Error accessing the database: {e}")
    finally:
        dbmain.close()

@app.route('/dashboard_feedback_reply/<int:feedback_id>', methods=['GET', 'POST'])
def dashboard_feedback_reply(feedback_id):
    dbmain = shelve.open('main.db', 'c')
    feedbacks_dict = dbmain.get('Feedback', {})
    feedback = feedbacks_dict.get(feedback_id)

    if not feedback:
        return render_template('error_page.html', message="Feedback not found.")

    if request.method == 'POST':
        reply1 = request.form.get('reply')
        reply = f"Reply:{reply1}"# Get reply text from the form
        feedback.set_reply(reply)  # Assume `set_reply` is a method in your Feedback class
        dbmain['Feedback'] = feedbacks_dict  # Save the updated feedback back to the database
        dbmain.close()
        return redirect(url_for('some_dashboard_page'))  # Redirect to the relevant dashboard

    dbmain.close()
    return render_template('feedback_reply.html', feedback=feedback)

#modifes delivery status , opstat
@app.route('/dashboard/transactions', methods=['GET', 'POST'])
def dashboard_transactions():
    searchform = SearchTransactionField(request.form)
    filterform = FilterTransactions(request.form)
    delivery_id = request.form.get('Delivery_id')
    new_status = request.form.get('new_status')
    
    dbmain = shelve.open('main.db', 'c')
    deliveries_dict = dbmain.get("delivery", {})


    try:
        if "delivery" in dbmain:
            deliveries_dict = dbmain["delivery"]  # sync local with db2
        else:
            dbmain['delivery'] = deliveries_dict  # sync db2 with local (basically null)
    except:
        print("Error in opening main.db")

    deliveries_list = list(deliveries_dict.values())

    if delivery_id:
        try:
            delivery_id = int(delivery_id)
        except ValueError:
            print("Invalid delivery ID format")
            delivery_id = None


    #opstats code here - in theory it shld work
    try:
        print(f"New status is {new_status}")
        deliveryobj = deliveries_dict.get(delivery_id)
        if deliveryobj.get_status() == "Pending":
            Operatorstats.operatorstats_feedbacks("Pending","minus")
            Operatorstats.operatorstats_feedbacks(new_status,"plus")
        elif deliveryobj.get_status() == "In Transit":
            Operatorstats.operatorstats_feedbacks("In Transit","minus")
            Operatorstats.operatorstats_feedbacks(new_status,"plus")
        elif deliveryobj.get_status() == "Delivered":
            Operatorstats.operatorstats_feedbacks("Delivered","minus")
            Operatorstats.operatorstats_feedbacks(new_status,"plus")
        else:
            pass
    except:
        pass
    
    if delivery_id and delivery_id in deliveries_dict:
        delivery = deliveries_dict[delivery_id]
        print(f"Updating Delivery ID {delivery_id} to status: {new_status}")
        delivery.set_status(new_status)  # Assuming a setter method exists
        dbmain['delivery'] = deliveries_dict
    elif delivery_id:
        print(f"Delivery ID {delivery_id} not found!")
    dbmain.close()


    if request.method == 'POST' and searchform.validate():
        return redirect(url_for('dashboardtransactionssearch',  keyword=searchform.searchfield.data))
    
    if request.method == 'POST' and filterform.validate():
        return redirect(url_for('dashboardtransactionsfilter',  keyword=filterform.searchstatusfield.data))
    
    return render_template('Operatordashboard_transaction.html',searchform=searchform , deliveries_list=deliveries_list,filterform = filterform)


@app.route('/dashboard/transactions/filter=<keyword>',methods=['GET','POST'])
def dashboardtransactionsfilter(keyword):
    if not keyword:
        return redirect(url_for('dashboard_transactions'))
    filterform = FilterTransactions(request.form)
    searchform = SearchTransactionField(request.form)
    dbmain = shelve.open('main.db', 'c')
    deliveries_dict = {}
    try:
        if "delivery" in dbmain:
            deliveries_dict = dbmain["delivery"]  # sync local with db2
        else:
            dbmain['delivery'] = deliveries_dict  # sync db2 with local (basically null)
    except:
        print("Error in opening main.db")
    deliveries_list = []
    for key in deliveries_dict:
        delivery = deliveries_dict.get(key)
        if delivery.get_status() == keyword:
            deliveries_list.append(delivery)
        else:
            pass

    if request.method == 'POST' and searchform.validate():
        return redirect(url_for('dashboardtransactionssearch', keyword=searchform.searchfield.data))

    if request.method == 'POST' and filterform.validate():
        return redirect(url_for('dashboardtransactionsfilter',  keyword=filterform.searchstatusfield.data))   
    
    return render_template('Operatordashboard_transaction_filter.html', searchform = searchform,
                           deliveries_list=deliveries_list,filterform = filterform, searchcondition = keyword)


@app.route('/dashboard/transactions/search/<keyword>',  methods=['GET', 'POST'])
def dashboardtransactionssearch(keyword):
    if not keyword:
        return redirect(url_for('dashboard_transactions'))
    filterform = FilterTransactions(request.form)
    searchform = SearchTransactionField(request.form)
    dbmain = shelve.open('main.db', 'c')
    deliveries_dict = {}

    try:
        if "delivery" in dbmain:
            deliveries_dict = dbmain["delivery"]  # sync local with db2
        else:
            dbmain['delivery'] = deliveries_dict  # sync db2 with local (basically null)
    except:
        print("Error in opening main.db")
    deliveries_list = []
    for key in deliveries_dict:
        delivery = deliveries_dict.get(key)
        if keyword in delivery.get_item_title():  # Assuming you're checking by item title
            deliveries_list.append(delivery)

    if request.method == 'POST' and searchform.validate():
        return redirect(url_for('dashboardtransactionssearch', keyword=searchform.searchfield.data))

    if request.method == 'POST' and filterform.validate():
        return redirect(url_for('dashboardtransactionsfilter',  keyword=filterform.searchstatusfield.data))

    return render_template('Operatordashboard_transaction_search.html', searchform = searchform,
                           deliveries_list=deliveries_list,filterform = filterform, searchcondition = keyword)

if __name__ == "__main__":
    
    app.secret_key = 'super secret key'
    app.config['SESSION_TYPE'] = 'filesystem'
    #determine the create the obj
    app.run(debug=True)
    
